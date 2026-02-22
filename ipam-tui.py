#!/usr/bin/env python3
import curses
import hashlib
import hmac
import io
import ipaddress
import re
import secrets
import shutil
import sqlite3
import sys
import time
import zlib
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

# =============================================================================
# Config
# =============================================================================

VERSION = "0.9.0"
APP_TITLE = "IPAM / VLAN Manager"
MAX_ENUM_HOSTS = 4096  # guard rail for enumerating "unused" IPs in a subnet
VLAN_SUBNET_KEYS = ["Customer", "Location", "Comment"]
STANDARD_KEYS = ["Customer", "Location", "Comment", "Asset", "Interface", "Network Connection"]

# User roles
ROLE_ADMIN = "admin"
ROLE_EDITOR = "editor"
ROLE_VIEWER = "viewer"

# Color pair IDs
CP_NORMAL = 1
CP_HILITE = 2
CP_BAR = 3
CP_DIM = 4

# =============================================================================
# Navigation control
# =============================================================================


class GoHome(Exception):
    """Raised when user presses ESC to jump back to main menu."""


# =============================================================================
# Database schema
# =============================================================================

SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS config (
  key TEXT PRIMARY KEY,
  value TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS users (
  id INTEGER PRIMARY KEY,
  username TEXT UNIQUE NOT NULL,
  password_hash TEXT NOT NULL,
  role TEXT NOT NULL DEFAULT 'viewer',
  fg_color TEXT DEFAULT 'green'
);

CREATE TABLE IF NOT EXISTS vlans (
  id INTEGER PRIMARY KEY,
  vlan_num INTEGER NOT NULL UNIQUE,
  name TEXT DEFAULT '',
  routed INTEGER NOT NULL DEFAULT 0,
  uplink TEXT DEFAULT ''
);

CREATE TABLE IF NOT EXISTS broadcast_domains (
  id INTEGER PRIMARY KEY,
  vlan_id INTEGER NOT NULL REFERENCES vlans(id) ON DELETE CASCADE,
  name TEXT NOT NULL,
  UNIQUE(vlan_id, name)
);

CREATE TABLE IF NOT EXISTS bd_ranges (
  id INTEGER PRIMARY KEY,
  bd_id INTEGER NOT NULL REFERENCES broadcast_domains(id) ON DELETE CASCADE,
  cidr TEXT NOT NULL,
  UNIQUE(bd_id, cidr)
);

CREATE TABLE IF NOT EXISTS ip_addresses (
  id INTEGER PRIMARY KEY,
  addr TEXT NOT NULL UNIQUE,
  vlan_id INTEGER REFERENCES vlans(id) ON DELETE SET NULL,
  bd_id INTEGER REFERENCES broadcast_domains(id) ON DELETE SET NULL
);

CREATE TABLE IF NOT EXISTS attributes (
  id INTEGER PRIMARY KEY,
  scope TEXT NOT NULL, -- 'vlan' | 'bd' | 'ip'
  scope_id INTEGER NOT NULL,
  key TEXT NOT NULL,
  value TEXT NOT NULL,
  inheritable INTEGER NOT NULL DEFAULT 0,
  UNIQUE(scope, scope_id, key)
);

CREATE INDEX IF NOT EXISTS idx_attr_key ON attributes(key);
CREATE INDEX IF NOT EXISTS idx_attr_scope ON attributes(scope, scope_id);
CREATE INDEX IF NOT EXISTS idx_attr_lookup ON attributes(scope, scope_id, key);

CREATE INDEX IF NOT EXISTS idx_ip_vlan ON ip_addresses(vlan_id);
CREATE INDEX IF NOT EXISTS idx_ip_bd ON ip_addresses(bd_id);
CREATE INDEX IF NOT EXISTS idx_bd_vlan ON broadcast_domains(vlan_id);

CREATE TABLE IF NOT EXISTS audit_log (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  timestamp TEXT NOT NULL,
  action TEXT NOT NULL,
  description TEXT NOT NULL,
  snapshot_id INTEGER,
  logged_by TEXT
);

CREATE TABLE IF NOT EXISTS snapshots (
  id INTEGER PRIMARY KEY AUTOINCREMENT,
  timestamp TEXT NOT NULL,
  db_state BLOB NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_audit_timestamp ON audit_log(timestamp DESC);
"""


@dataclass
class Resolved:
    vlan_id: Optional[int]
    bd_id: Optional[int]
    vlan_num: Optional[int]
    vlan_name: Optional[str]
    bd_name: Optional[str]


@dataclass
class ListRow:
    label: str
    customer: str
    location: str


@dataclass
class User:
    id: int
    username: str
    role: str
    fg_color: str

    def is_admin(self) -> bool:
        return self.role == ROLE_ADMIN

    def is_viewer(self) -> bool:
        return self.role == ROLE_VIEWER

    def can_edit(self) -> bool:
        return self.role in (ROLE_ADMIN, ROLE_EDITOR)


def collapse_set(s: set) -> str:
    """Collapse a set of values into a single display string."""
    if len(s) == 0:
        return ""
    if len(s) == 1:
        return next(iter(s))
    return "Multiple"


class DB:
    def __init__(self, path: str):
        self.con = sqlite3.connect(path, check_same_thread=False)
        self.con.row_factory = sqlite3.Row
        self._in_transaction = False
        self.current_user: Optional[User] = None
        self.db_name: str = ""
        self._resolve_cache: Optional[list] = None  # Cached parsed bd_ranges for resolve_for_ip

    def invalidate_resolve_cache(self):
        """Clear the bd_ranges cache. Call after any change to ranges, subnets, or VLANs."""
        self._resolve_cache = None

    def init(self):
        self.con.executescript(SCHEMA_SQL)
        self.x("INSERT OR IGNORE INTO config(key,value) VALUES('db_name','')")
        self.x("INSERT OR IGNORE INTO config(key,value) VALUES('snapshot_max_count','20')")
        self.x("INSERT OR IGNORE INTO config(key,value) VALUES('snapshot_enabled','1')")

        # Create default admin user if no users exist
        user_count = self.q("SELECT COUNT(*) as cnt FROM users")[0]["cnt"]
        if user_count == 0:
            default_hash = self.hash_password("admin")
            self.x(
                "INSERT INTO users(username, password_hash, role, fg_color) VALUES(?, ?, ?, ?)",
                ("admin", default_hash, ROLE_ADMIN, "green")
            )

        # Migration: add logged_by column if it doesn't exist (for upgrades from v1.x)
        try:
            self.q("SELECT logged_by FROM audit_log LIMIT 1")
        except sqlite3.OperationalError:
            self.x("ALTER TABLE audit_log ADD COLUMN logged_by TEXT")

        self.con.commit()

    def close(self):
        self.con.close()

    def set_current_user(self, user: Optional['User']):
        self.current_user = user

    def transaction(self):
        """Context manager for manual transaction control."""

        @contextmanager
        def _transaction():
            if self._in_transaction:
                yield  # Nested transaction, just yield
                return

            self._in_transaction = True
            try:
                yield
                self.con.commit()
            except Exception:
                self.con.rollback()
                raise
            finally:
                self._in_transaction = False

        return _transaction()

    def q(self, sql: str, params: Tuple = ()) -> List[sqlite3.Row]:
        return self.con.execute(sql, params).fetchall()

    def x(self, sql: str, params: Tuple = ()) -> int:
        cur = self.con.execute(sql, params)
        if not self._in_transaction:
            self.con.commit()
        return cur.lastrowid

    # ---- Config ----

    def get_config(self, key: str, default: str = "") -> str:
        rows = self.q("SELECT value FROM config WHERE key=?", (key,))
        if not rows:
            return default
        return (rows[0]["value"] or "").strip()

    def set_config(self, key: str, value: str):
        self.x(
            """
            INSERT INTO config(key,value) VALUES(?,?)
            ON CONFLICT(key) DO UPDATE SET value=excluded.value
            """,
            (key, value),
        )

    # ---- Audit Log & Snapshots ----

    def create_snapshot(self) -> int:
        """Create a snapshot of the current database state. Returns snapshot_id."""
        # Create an in-memory backup of the database
        backup_data = io.BytesIO()
        backup_con = sqlite3.connect(':memory:')

        # Copy all tables except audit_log and snapshots
        for line in self.con.iterdump():
            if 'audit_log' not in line and 'snapshots' not in line and 'sqlite_sequence' not in line:
                backup_con.execute(line)
        backup_con.commit()

        # Serialize to bytes
        for line in backup_con.iterdump():
            backup_data.write(line.encode('utf-8'))
            backup_data.write(b'\n')

        backup_con.close()
        db_state = backup_data.getvalue()

        # Compress the snapshot
        compressed_state = zlib.compress(db_state, level=9)

        timestamp = datetime.now().isoformat()
        snapshot_id = self.x(
            "INSERT INTO snapshots(timestamp, db_state) VALUES(?, ?)",
            (timestamp, compressed_state)
        )

        # Cleanup old snapshots
        self.cleanup_old_snapshots()

        return snapshot_id

    def log_action(self, action: str, description: str, create_snapshot: bool = False):
        """Log an action to the audit log, optionally creating a snapshot."""
        snapshot_id = None
        if create_snapshot:
            # Check if snapshots are enabled
            snapshot_enabled = self.get_config("snapshot_enabled", "1") == "1"
            if snapshot_enabled:
                snapshot_id = self.create_snapshot()

        timestamp = datetime.now().isoformat()
        logged_by = self.current_user.username if self.current_user else None
        self.x(
            "INSERT INTO audit_log(timestamp, action, description, snapshot_id, logged_by) VALUES(?, ?, ?, ?, ?)",
            (timestamp, action, description, snapshot_id, logged_by)
        )

    def list_audit_log(self, limit: Optional[int] = None) -> List[sqlite3.Row]:
        """Get audit log entries, optionally limited."""
        if limit:
            return self.q(
                "SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT ?",
                (limit,)
            )
        return self.q("SELECT * FROM audit_log ORDER BY timestamp DESC")

    def cleanup_old_snapshots(self):
        """Remove snapshots beyond retention policy."""
        max_snapshots = int(self.get_config("snapshot_max_count", "20"))

        # Get all snapshots ordered by timestamp descending
        all_snapshots = self.q("SELECT id, timestamp FROM snapshots ORDER BY timestamp DESC")

        if len(all_snapshots) <= max_snapshots:
            return  # Under limit, no cleanup needed

        # Retention policy:
        # - Keep last 10 unconditionally
        # - Keep 1 per day for last 7 days
        # - Keep 1 per week for last 4 weeks
        # - Delete everything else

        keep_ids = set()
        now = datetime.now()

        # Keep last 10 unconditionally
        for snapshot in all_snapshots[:10]:
            keep_ids.add(snapshot["id"])

        # Group remaining by date buckets
        daily_buckets = {}  # date -> snapshot_id
        weekly_buckets = {}  # week_number -> snapshot_id

        for snapshot in all_snapshots[10:]:
            try:
                snap_time = datetime.fromisoformat(snapshot["timestamp"])
                age_days = (now - snap_time).days

                # Keep 1 per day for last 7 days
                if age_days <= 7:
                    date_key = snap_time.date()
                    if date_key not in daily_buckets:
                        daily_buckets[date_key] = snapshot["id"]
                        keep_ids.add(snapshot["id"])

                # Keep 1 per week for last 4 weeks
                elif age_days <= 28:
                    week_key = snap_time.isocalendar()[1]  # ISO week number
                    if week_key not in weekly_buckets:
                        weekly_buckets[week_key] = snapshot["id"]
                        keep_ids.add(snapshot["id"])
            except (ValueError, KeyError):
                # If timestamp parsing fails, keep it to be safe
                keep_ids.add(snapshot["id"])

        # Delete snapshots not in keep list
        for snapshot in all_snapshots:
            if snapshot["id"] not in keep_ids:
                self.x("DELETE FROM snapshots WHERE id=?", (snapshot["id"],))

    def get_snapshot_stats(self) -> Dict[str, Any]:
        """Get snapshot statistics."""
        stats = {
            "count": 0,
            "total_size_mb": 0.0,
            "oldest_date": None,
            "newest_date": None,
        }

        rows = self.q("SELECT COUNT(*) as cnt, SUM(LENGTH(db_state)) as total FROM snapshots")
        if rows and rows[0]["cnt"]:
            stats["count"] = rows[0]["cnt"]
            stats["total_size_mb"] = (rows[0]["total"] or 0) / (1024 * 1024)

        oldest = self.q("SELECT timestamp FROM snapshots ORDER BY timestamp ASC LIMIT 1")
        if oldest:
            stats["oldest_date"] = oldest[0]["timestamp"][:19]

        newest = self.q("SELECT timestamp FROM snapshots ORDER BY timestamp DESC LIMIT 1")
        if newest:
            stats["newest_date"] = newest[0]["timestamp"][:19]

        return stats

    def get_snapshot(self, snapshot_id: int) -> Optional[bytes]:
        """Retrieve a snapshot by ID."""
        rows = self.q("SELECT db_state FROM snapshots WHERE id=?", (snapshot_id,))
        return rows[0]["db_state"] if rows else None

    def restore_snapshot(self, snapshot_id: int, backup_path: str):
        """Restore database to a snapshot state. Creates backup first.
        Preserves snapshots and audit_log tables so further rollbacks are possible."""
        # Create backup of current database
        db_path = self.con.execute("PRAGMA database_list").fetchone()[2]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f"{backup_path}.backup_{timestamp}"

        # Fetch snapshot data before closing connection
        rows = self.q("SELECT db_state FROM snapshots WHERE id=?", (snapshot_id,))
        if not rows:
            raise ValueError(f"Snapshot {snapshot_id} not found")
        snapshot_data = rows[0]["db_state"]

        # Decompress snapshot data
        try:
            decompressed_data = zlib.decompress(snapshot_data)
            snapshot_sql = decompressed_data.decode('utf-8')
        except zlib.error:
            # Assume it's uncompressed (legacy snapshot)
            snapshot_sql = snapshot_data.decode('utf-8')

        # Validate that we have actual SQL to restore
        if not snapshot_sql.strip():
            raise ValueError("Snapshot data is empty")

        # Preserve current snapshots and audit_log tables
        preserved_snapshots = self.q("SELECT id, timestamp, db_state FROM snapshots")
        preserved_audit = self.q("SELECT id, timestamp, action, description, snapshot_id FROM audit_log")

        # Close connection temporarily
        self.con.close()

        # Copy current database to backup
        shutil.copy2(db_path, backup_file)

        # Clear database and restore
        restore_con = sqlite3.connect(db_path)
        restore_con.executescript("PRAGMA writable_schema = 1; DELETE FROM sqlite_master; PRAGMA writable_schema = 0; VACUUM;")
        restore_con.executescript(snapshot_sql)
        restore_con.commit()

        # Restore the snapshots and audit_log tables
        restore_con.executescript("""
            CREATE TABLE IF NOT EXISTS audit_log (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              timestamp TEXT NOT NULL,
              action TEXT NOT NULL,
              description TEXT NOT NULL,
              snapshot_id INTEGER
            );
            CREATE TABLE IF NOT EXISTS snapshots (
              id INTEGER PRIMARY KEY AUTOINCREMENT,
              timestamp TEXT NOT NULL,
              db_state BLOB NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_audit_timestamp ON audit_log(timestamp DESC);
        """)

        # Re-insert preserved data
        for row in preserved_snapshots:
            restore_con.execute(
                "INSERT OR REPLACE INTO snapshots(id, timestamp, db_state) VALUES(?, ?, ?)",
                (row["id"], row["timestamp"], row["db_state"])
            )
        for row in preserved_audit:
            restore_con.execute(
                "INSERT OR REPLACE INTO audit_log(id, timestamp, action, description, snapshot_id) VALUES(?, ?, ?, ?, ?)",
                (row["id"], row["timestamp"], row["action"], row["description"], row["snapshot_id"])
            )

        # Log the restore action
        restore_con.execute(
            "INSERT INTO audit_log(timestamp, action, description, snapshot_id) VALUES(?, ?, ?, ?)",
            (datetime.now().isoformat(), "restore_snapshot", f"Restored to snapshot {snapshot_id}", None)
        )

        restore_con.commit()
        restore_con.close()

        # Reconnect
        self.con = sqlite3.connect(db_path, check_same_thread=False)
        self.con.row_factory = sqlite3.Row
        self.invalidate_resolve_cache()

        return backup_file

    # ---- VLAN ----

    def create_vlan(self, vlan_num: int, name: str, routed: int = 0, uplink: str = "") -> int:
        if not (1 <= vlan_num <= 4094):
            raise ValueError(f"VLAN number must be between 1 and 4094, got {vlan_num}")
        vlan_id = self.x("INSERT INTO vlans(vlan_num, name, routed, uplink) VALUES(?,?,?,?)", (vlan_num, name, routed, uplink))
        self.log_action("create_vlan", f"Created VLAN {vlan_num}: {name}", create_snapshot=True)
        return vlan_id

    def vlan_exists(self, vlan_num: int) -> bool:
        return bool(self.q("SELECT 1 FROM vlans WHERE vlan_num=?", (vlan_num,)))

    def get_vlan_by_num(self, vlan_num: int) -> Optional[sqlite3.Row]:
        rows = self.q("SELECT * FROM vlans WHERE vlan_num=?", (vlan_num,))
        return rows[0] if rows else None

    def get_vlan_by_id(self, vlan_id: int) -> Optional[sqlite3.Row]:
        rows = self.q("SELECT * FROM vlans WHERE id=?", (vlan_id,))
        return rows[0] if rows else None

    def list_vlans(self) -> List[sqlite3.Row]:
        return self.q("SELECT * FROM vlans ORDER BY vlan_num ASC")

    def update_vlan_routed(self, vlan_id: int, routed: int):
        self.x("UPDATE vlans SET routed=? WHERE id=?", (routed, vlan_id))

    def update_vlan_uplink(self, vlan_id: int, uplink: str):
        self.x("UPDATE vlans SET uplink=? WHERE id=?", (uplink, vlan_id))

    def delete_vlan(self, vlan_id: int) -> Tuple[int, int]:
        """Delete a VLAN and all its child subnets and IPs.
        Returns (subnet_count, ip_count) of deleted items."""
        vlan = self.get_vlan_by_id(vlan_id)
        vlan_num = vlan["vlan_num"] if vlan else "?"
        vlan_name = vlan["name"] if vlan else "?"

        # Count what will be deleted
        subnets = self.list_subnets(vlan_id)
        subnet_count = len(subnets)
        ip_count = 0
        for subnet in subnets:
            ip_rows = self.list_ip_rows_in_subnet(subnet["id"])
            ip_count += len(ip_rows)

        # Log before deletion (with snapshot)
        self.log_action(
            "delete_vlan",
            f"Deleted VLAN {vlan_num} '{vlan_name}' with {subnet_count} subnet(s) and {ip_count} IP(s)",
            create_snapshot=True
        )

        # Delete VLAN (CASCADE will handle subnets and ranges, but we need to clean up IPs and attributes)
        # First, delete IP attributes and IP addresses linked to this VLAN's subnets
        for subnet in subnets:
            ip_rows = self.list_ip_rows_in_subnet(subnet["id"])
            if ip_rows:
                ip_ids = [r["id"] for r in ip_rows]
                placeholders = ','.join(['?'] * len(ip_ids))
                self.x(f"DELETE FROM attributes WHERE scope='ip' AND scope_id IN ({placeholders})", tuple(ip_ids))
            self.x("DELETE FROM ip_addresses WHERE bd_id=?", (subnet["id"],))
            self.x("DELETE FROM attributes WHERE scope='bd' AND scope_id=?", (subnet["id"],))

        # Delete VLAN attributes
        self.x("DELETE FROM attributes WHERE scope='vlan' AND scope_id=?", (vlan_id,))

        # Delete the VLAN (CASCADE handles broadcast_domains and bd_ranges)
        self.x("DELETE FROM vlans WHERE id=?", (vlan_id,))
        self.invalidate_resolve_cache()

        return subnet_count, ip_count

    def check_routed_vlan_overlap(self, vlan_id: int, new_cidr: str) -> Optional[Tuple[int, str]]:
        """Check if new_cidr overlaps with subnets in other routed VLANs.
        Returns (vlan_num, existing_cidr) if overlap found, None otherwise."""
        new_net = ipaddress.ip_network(new_cidr, strict=False)

        # Get all subnet ranges from other routed VLANs
        rows = self.q(
            """
            SELECT v.vlan_num, br.cidr
            FROM bd_ranges br
            JOIN broadcast_domains bd ON bd.id = br.bd_id
            JOIN vlans v ON v.id = bd.vlan_id
            WHERE v.routed = 1 AND v.id != ?
            """,
            (vlan_id,)
        )

        for r in rows:
            try:
                existing_net = ipaddress.ip_network(r["cidr"], strict=False)
                if new_net.overlaps(existing_net):
                    return (r["vlan_num"], r["cidr"])
            except ValueError:
                continue

        return None

    def check_vlan_can_be_routed(self, vlan_id: int) -> Optional[Tuple[str, int, str]]:
        """Check if a VLAN's subnets would conflict if marked as routed.
        Returns (this_cidr, other_vlan_num, other_cidr) if conflict found, None otherwise."""
        # Get all subnets in this VLAN
        my_ranges = self.q(
            """
            SELECT br.cidr
            FROM bd_ranges br
            JOIN broadcast_domains bd ON bd.id = br.bd_id
            WHERE bd.vlan_id = ?
            """,
            (vlan_id,)
        )

        # Check each against other routed VLANs
        for my_r in my_ranges:
            try:
                my_net = ipaddress.ip_network(my_r["cidr"], strict=False)
            except ValueError:
                continue

            # Get all ranges from other routed VLANs
            other_rows = self.q(
                """
                SELECT v.vlan_num, br.cidr
                FROM bd_ranges br
                JOIN broadcast_domains bd ON bd.id = br.bd_id
                JOIN vlans v ON v.id = bd.vlan_id
                WHERE v.routed = 1 AND v.id != ?
                """,
                (vlan_id,)
            )

            for other_r in other_rows:
                try:
                    other_net = ipaddress.ip_network(other_r["cidr"], strict=False)
                    if my_net.overlaps(other_net):
                        return (my_r["cidr"], other_r["vlan_num"], other_r["cidr"])
                except ValueError:
                    continue

        return None

    # ---- Subnet (broadcast domain) ----

    def create_subnet(self, vlan_id: int, name: str) -> int:
        bd_id = self.x("INSERT INTO broadcast_domains(vlan_id,name) VALUES(?,?)", (vlan_id, name))
        vlan = self.get_vlan_by_id(vlan_id)
        vlan_num = vlan["vlan_num"] if vlan else "?"
        self.log_action("create_subnet", f"Created subnet '{name}' in VLAN {vlan_num}", create_snapshot=True)
        return bd_id

    def list_subnets(self, vlan_id: Optional[int] = None) -> List[sqlite3.Row]:
        if vlan_id is None:
            return self.q(
                """
                SELECT bd.*, v.vlan_num AS vlan_num, v.name AS vlan_name
                FROM broadcast_domains bd
                JOIN vlans v ON v.id=bd.vlan_id
                ORDER BY v.vlan_num, bd.name
                """
            )
        return self.q("SELECT * FROM broadcast_domains WHERE vlan_id=? ORDER BY name", (vlan_id,))

    def list_subnets_with_ranges(self) -> List[sqlite3.Row]:
        return self.q(
            """
            SELECT bd.*, v.vlan_num AS vlan_num, v.name AS vlan_name, COUNT(br.id) AS range_count
            FROM broadcast_domains bd
            JOIN vlans v ON v.id=bd.vlan_id
            LEFT JOIN bd_ranges br ON br.bd_id=bd.id
            GROUP BY bd.id
            HAVING COUNT(br.id) > 0
            ORDER BY v.vlan_num, bd.name
            """
        )

    def get_subnet(self, bd_id: int) -> Optional[sqlite3.Row]:
        rows = self.q("SELECT * FROM broadcast_domains WHERE id=?", (bd_id,))
        return rows[0] if rows else None

    def add_subnet_range(self, bd_id: int, cidr: str) -> bool:
        """Add a CIDR range to a subnet. Returns True if added, False if duplicate."""
        net = ipaddress.ip_network(cidr, strict=False)
        normalized_cidr = str(net)
        existing = self.q("SELECT 1 FROM bd_ranges WHERE bd_id=? AND cidr=?", (bd_id, normalized_cidr))
        if existing:
            return False

        # Check for overlaps with existing ranges in this same subnet
        sibling_ranges = self.q("SELECT cidr FROM bd_ranges WHERE bd_id=?", (bd_id,))
        for r in sibling_ranges:
            try:
                existing_net = ipaddress.ip_network(r["cidr"], strict=False)
            except ValueError:
                continue
            if net.overlaps(existing_net):
                raise ValueError(f"Range {normalized_cidr} overlaps with existing range {r['cidr']} in this subnet")

        # Check if this subnet's VLAN is routed, and if so, validate no overlaps
        subnet = self.get_subnet(bd_id)
        if subnet:
            vlan = self.get_vlan_by_id(subnet["vlan_id"])
            if vlan and vlan["routed"]:
                overlap = self.check_routed_vlan_overlap(subnet["vlan_id"], normalized_cidr)
                if overlap:
                    raise ValueError(f"Subnet overlaps with VLAN {overlap[0]} subnet {overlap[1]} (both VLANs are routed)")

        self.x("INSERT INTO bd_ranges(bd_id,cidr) VALUES(?,?)", (bd_id, normalized_cidr))
        self.invalidate_resolve_cache()
        return True

    def list_subnet_ranges(self, bd_id: int) -> List[sqlite3.Row]:
        return self.q("SELECT * FROM bd_ranges WHERE bd_id=? ORDER BY cidr", (bd_id,))

    def move_subnet_to_vlan(self, bd_id: int, new_vlan_id: int):
        self.x("UPDATE broadcast_domains SET vlan_id=? WHERE id=?", (new_vlan_id, bd_id))
        self.invalidate_resolve_cache()

    def update_subnet_name(self, bd_id: int, name: str):
        self.x("UPDATE broadcast_domains SET name=? WHERE id=?", (name, bd_id))

    def delete_subnet(self, bd_id: int):
        subnet = self.get_subnet(bd_id)
        name = subnet["name"] if subnet else "?"
        self.log_action("delete_subnet", f"Deleted subnet '{name}'", create_snapshot=True)
        # Clean up IP attributes and IP records before deleting the subnet
        ip_rows = self.list_ip_rows_in_subnet(bd_id)
        if ip_rows:
            ip_ids = [r["id"] for r in ip_rows]
            placeholders = ','.join(['?'] * len(ip_ids))
            self.x(f"DELETE FROM attributes WHERE scope='ip' AND scope_id IN ({placeholders})", tuple(ip_ids))
            self.x("DELETE FROM ip_addresses WHERE bd_id=?", (bd_id,))
        self.x("DELETE FROM attributes WHERE scope='bd' AND scope_id=?", (bd_id,))
        self.x("DELETE FROM broadcast_domains WHERE id=?", (bd_id,))
        self.invalidate_resolve_cache()

    # ---- Attributes ----

    def get_attrs(self, scope: str, scope_id: int) -> Dict[str, str]:
        rows = self.q("SELECT key,value FROM attributes WHERE scope=? AND scope_id=?", (scope, scope_id))
        out = {r["key"]: r["value"] for r in rows}
        # Default keys based on scope
        default_keys = VLAN_SUBNET_KEYS if scope in ('vlan', 'bd') else STANDARD_KEYS
        for k in default_keys:
            out.setdefault(k, "")
        return out

    def upsert_attr(self, scope: str, scope_id: int, key: str, value: str, inheritable: int):
        key = key.strip()
        self.x(
            """
            INSERT INTO attributes(scope,scope_id,key,value,inheritable)
            VALUES(?,?,?,?,?)
            ON CONFLICT(scope,scope_id,key) DO UPDATE SET value=excluded.value,inheritable=excluded.inheritable
            """,
            (scope, scope_id, key, value, inheritable),
        )
        self.log_action("update_attribute", f"Updated {scope} attribute: {key} = {value}")

    def delete_attr(self, scope: str, scope_id: int, key: str):
        key = key.strip()
        self.x("DELETE FROM attributes WHERE scope=? AND scope_id=? AND key=?", (scope, scope_id, key))

    # ---- IPs ----

    def ensure_ip(self, ip_str: str) -> int:
        # Use transaction to prevent race condition between INSERT and SELECT
        ip = str(ipaddress.ip_address(ip_str))
        with self.transaction():
            self.x("INSERT OR IGNORE INTO ip_addresses(addr) VALUES(?)", (ip,))
            rows = self.q("SELECT id FROM ip_addresses WHERE addr=?", (ip,))
        return rows[0]["id"]

    def get_ip_row(self, ip_str: str) -> Optional[sqlite3.Row]:
        ip = str(ipaddress.ip_address(ip_str))
        rows = self.q("SELECT * FROM ip_addresses WHERE addr=?", (ip,))
        return rows[0] if rows else None

    def set_ip_links(self, ip_id: int, vlan_id: Optional[int], bd_id: Optional[int]):
        self.x("UPDATE ip_addresses SET vlan_id=?, bd_id=? WHERE id=?", (vlan_id, bd_id, ip_id))

    def list_ip_rows_in_subnet(self, bd_id: int) -> List[sqlite3.Row]:
        return self.q("SELECT * FROM ip_addresses WHERE bd_id=? ORDER BY addr", (bd_id,))

    # ---- Resolve / effective ----

    def resolve_for_ip(self, ip_str: str) -> Resolved:
        ip = ipaddress.ip_address(ip_str)

        # Build / reuse cached parsed bd_ranges
        if self._resolve_cache is None:
            bd_rows = self.q(
                """
                SELECT br.cidr,
                       bd.id AS bd_id, bd.name AS bd_name,
                       v.id AS vlan_id, v.vlan_num AS vlan_num, v.name AS vlan_name
                FROM bd_ranges br
                JOIN broadcast_domains bd ON bd.id = br.bd_id
                JOIN vlans v ON v.id = bd.vlan_id
                """
            )
            parsed = []
            for r in bd_rows:
                try:
                    net = ipaddress.ip_network(r["cidr"], strict=False)
                except ValueError:
                    continue
                parsed.append((net, r))
            self._resolve_cache = parsed

        best = None  # (prefixlen, row)
        for net, r in self._resolve_cache:
            if ip in net:
                pl = net.prefixlen
                if best is None or pl > best[0]:
                    best = (pl, r)

        if best is not None:
            r = best[1]
            return Resolved(
                vlan_id=r["vlan_id"],
                bd_id=r["bd_id"],
                vlan_num=r["vlan_num"],
                vlan_name=r["vlan_name"],
                bd_name=r["bd_name"],
            )

        # No match found
        return Resolved(vlan_id=None, bd_id=None, vlan_num=None, vlan_name=None, bd_name=None)


    def inherited_attrs_for_ip(self, ip_str: str) -> Tuple[Resolved, Dict[str, str]]:
        res = self.resolve_for_ip(ip_str)
        inherited: Dict[str, str] = {}
        if res.vlan_id is not None:
            inherited.update(self.get_attrs("vlan", res.vlan_id))
        if res.bd_id is not None:
            inherited.update(self.get_attrs("bd", res.bd_id))
        for k in STANDARD_KEYS:
            inherited.setdefault(k, "")
        return res, inherited

    def effective_attrs_for_ip(self, ip_str: str) -> Tuple[Resolved, Dict[str, str]]:
        res, inherited = self.inherited_attrs_for_ip(ip_str)
        eff = dict(inherited)
        ip_row = self.get_ip_row(ip_str)
        if ip_row:
            ip_attrs = self.get_attrs("ip", ip_row["id"])
            # Only apply non-empty IP attributes, allowing parent values to show through when cleared
            for k, v in ip_attrs.items():
                if v.strip():  # Only override inherited value if IP attribute is non-empty
                    eff[k] = v
        for k in STANDARD_KEYS:
            eff.setdefault(k, "")
        return res, eff

    # ---- Aggregation for list columns ----

    def batch_aggregate_for_vlans(self, vlan_ids: List[int]) -> Dict[int, Tuple[str, str]]:
        """Batch load Customer/Location aggregates for multiple VLANs efficiently."""
        if not vlan_ids:
            return {}

        placeholders = ','.join(['?'] * len(vlan_ids))

        # Single query to get all relevant attributes
        rows = self.q(
            f"""
            SELECT vlan_id, key, value FROM (
                SELECT bd.vlan_id, a.key, a.value
                FROM broadcast_domains bd
                JOIN attributes a ON a.scope='bd' AND a.scope_id=bd.id
                WHERE bd.vlan_id IN ({placeholders}) AND a.key IN ('Customer', 'Location')
                UNION ALL
                SELECT ip.vlan_id, a.key, a.value
                FROM ip_addresses ip
                JOIN attributes a ON a.scope='ip' AND a.scope_id=ip.id
                WHERE ip.vlan_id IN ({placeholders}) AND a.key IN ('Customer', 'Location')
                UNION ALL
                SELECT ?, a.key, a.value FROM attributes a
                WHERE a.scope='vlan' AND a.scope_id IN ({placeholders}) AND a.key IN ('Customer', 'Location')
            )
            """,
            tuple(vlan_ids) + tuple(vlan_ids) + (0,) + tuple(vlan_ids),
        )

        # Also get direct VLAN attributes with correct vlan_id mapping
        vlan_attr_rows = self.q(
            f"""
            SELECT scope_id as vlan_id, key, value FROM attributes
            WHERE scope='vlan' AND scope_id IN ({placeholders}) AND key IN ('Customer', 'Location')
            """,
            tuple(vlan_ids),
        )

        # Aggregate by vlan_id
        data: Dict[int, Dict[str, set]] = {vid: {"Customer": set(), "Location": set()} for vid in vlan_ids}

        for r in rows:
            vid = r["vlan_id"]
            if vid in data:
                v = (r["value"] or "").strip()
                if v:
                    data[vid][r["key"]].add(v)

        for r in vlan_attr_rows:
            vid = r["vlan_id"]
            if vid in data:
                v = (r["value"] or "").strip()
                if v:
                    data[vid][r["key"]].add(v)

        return {vid: (collapse_set(d["Customer"]), collapse_set(d["Location"])) for vid, d in data.items()}

    def batch_aggregate_for_subnets(self, bd_ids: List[int]) -> Dict[int, Tuple[str, str]]:
        """Batch load Customer/Location aggregates for multiple subnets efficiently."""
        if not bd_ids:
            return {}

        placeholders = ','.join(['?'] * len(bd_ids))

        # Single query to get all relevant attributes
        rows = self.q(
            f"""
            SELECT bd_id, key, value FROM (
                SELECT a.scope_id as bd_id, a.key, a.value FROM attributes a
                WHERE a.scope='bd' AND a.scope_id IN ({placeholders}) AND a.key IN ('Customer', 'Location')
                UNION ALL
                SELECT ip.bd_id, a.key, a.value FROM ip_addresses ip
                JOIN attributes a ON a.scope='ip' AND a.scope_id=ip.id
                WHERE ip.bd_id IN ({placeholders}) AND a.key IN ('Customer', 'Location')
            )
            """,
            tuple(bd_ids) + tuple(bd_ids),
        )

        # Aggregate by bd_id
        data: Dict[int, Dict[str, set]] = {bid: {"Customer": set(), "Location": set()} for bid in bd_ids}

        for r in rows:
            bid = r["bd_id"]
            if bid in data:
                v = (r["value"] or "").strip()
                if v:
                    data[bid][r["key"]].add(v)

        return {bid: (collapse_set(d["Customer"]), collapse_set(d["Location"])) for bid, d in data.items()}

    def get_all_custom_keys(self) -> List[str]:
        """Get all custom attribute keys used anywhere in the database."""
        placeholders = ','.join(['?'] * len(STANDARD_KEYS))
        rows = self.q(f"SELECT DISTINCT key FROM attributes WHERE key NOT IN ({placeholders}) ORDER BY key", tuple(STANDARD_KEYS))
        return [r["key"] for r in rows]

    # ---- User Management ----

    # Password requirements hint shown on password fields
    def validate_username(self, username: str) -> Tuple[bool, str]:
        """Validate username meets requirements. Returns (valid, error_message)."""
        username = username.strip()
        if not username:
            return False, "Username cannot be empty"
        if len(username) < 2:
            return False, "Username must be at least 2 characters"
        if len(username) > 64:
            return False, "Username must be 64 characters or fewer"
        if not re.match(r'^[a-zA-Z0-9._-]+$', username):
            return False, "Username may only contain letters, numbers, dots, hyphens, and underscores"
        return True, ""

    PASSWORD_REQUIREMENTS = [
        "Min 16 characters",
        "Uppercase + lowercase",
        "Number + special char",
    ]

    # Blocked example passphrases (lowercase for comparison)
    BLOCKED_PASSPHRASES = [
        "tango4$emotional$bagpipes",
        "correct horse battery staple",
        "correcthorsebatterystaple",
    ]

    def validate_password(self, password: str) -> Tuple[bool, str]:
        """Validate password meets requirements. Returns (valid, error_message)."""
        if len(password) < 16:
            return False, f"Password must be at least 16 characters (currently {len(password)})"

        if not any(c.isupper() for c in password):
            return False, "Password must contain at least one uppercase letter (A-Z)"

        if not any(c.islower() for c in password):
            return False, "Password must contain at least one lowercase letter (a-z)"

        if not any(c.isdigit() for c in password):
            return False, "Password must contain at least one number (0-9)"

        special_chars = "!@#$%^&*()_+-=[]{}|;':\",./<>?`~"
        if not any(c in special_chars for c in password):
            return False, "Password must contain at least one special character (!@#$%^&* etc.)"

        # Check for blocked example passphrases
        password_lower = password.lower()
        for blocked in self.BLOCKED_PASSPHRASES:
            if blocked in password_lower:
                return False, "Please don't use the example passphrase - create your own!"

        return True, ""

    def hash_password(self, password: str) -> str:
        """Hash a password using PBKDF2-SHA256 with a random salt."""
        salt = secrets.token_hex(16)
        h = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 600_000).hex()
        return f"{salt}:{h}"

    def verify_password(self, password: str, stored_hash: str) -> bool:
        """Verify a password against a stored hash. Supports both PBKDF2 and legacy SHA-256."""
        if ':' in stored_hash:
            salt, h = stored_hash.split(':', 1)
            candidate = hashlib.pbkdf2_hmac('sha256', password.encode(), salt.encode(), 600_000).hex()
            return hmac.compare_digest(candidate, h)
        else:
            # Legacy SHA-256 (unsalted) — accept but caller should upgrade
            candidate = hashlib.sha256(password.encode()).hexdigest()
            return hmac.compare_digest(candidate, stored_hash)

    def authenticate(self, username: str, password: str) -> Optional[User]:
        """Authenticate a user. Returns User object if successful, None otherwise."""
        rows = self.q(
            "SELECT id, username, role, fg_color, password_hash FROM users WHERE username=?",
            (username,)
        )
        if not rows:
            return None
        r = rows[0]
        if not self.verify_password(password, r["password_hash"]):
            return None
        # Auto-upgrade legacy SHA-256 hashes to PBKDF2 on successful login
        if ':' not in r["password_hash"]:
            self.x("UPDATE users SET password_hash=? WHERE id=?", (self.hash_password(password), r["id"]))
        return User(id=r["id"], username=r["username"], role=r["role"], fg_color=r["fg_color"])

    def get_user_by_username(self, username: str) -> Optional[User]:
        """Get a user by username."""
        rows = self.q("SELECT id, username, role, fg_color FROM users WHERE username=?", (username,))
        if rows:
            r = rows[0]
            return User(id=r["id"], username=r["username"], role=r["role"], fg_color=r["fg_color"])
        return None

    def get_user_by_id(self, user_id: int) -> Optional[User]:
        """Get a user by ID."""
        rows = self.q("SELECT id, username, role, fg_color FROM users WHERE id=?", (user_id,))
        if rows:
            r = rows[0]
            return User(id=r["id"], username=r["username"], role=r["role"], fg_color=r["fg_color"])
        return None

    def list_users(self) -> List[User]:
        """List all users."""
        rows = self.q("SELECT id, username, role, fg_color FROM users ORDER BY username")
        return [User(id=r["id"], username=r["username"], role=r["role"], fg_color=r["fg_color"]) for r in rows]

    def create_user(self, username: str, password: str, role: str) -> int:
        """Create a new user. Returns user ID."""
        password_hash = self.hash_password(password)
        user_id = self.x(
            "INSERT INTO users(username, password_hash, role, fg_color) VALUES(?, ?, ?, ?)",
            (username, password_hash, role, "green")
        )
        self.log_action("create_user", f"Created user '{username}' with role '{role}'")
        return user_id

    def update_user_password(self, user_id: int, new_password: str):
        """Update a user's password."""
        password_hash = self.hash_password(new_password)
        self.x("UPDATE users SET password_hash=? WHERE id=?", (password_hash, user_id))
        user = self.get_user_by_id(user_id)
        username = user.username if user else f"id={user_id}"
        self.log_action("change_password", f"Password changed for user '{username}'")

    def update_user_role(self, user_id: int, new_role: str):
        """Update a user's role."""
        user = self.get_user_by_id(user_id)
        old_role = user.role if user else "unknown"
        username = user.username if user else f"id={user_id}"
        self.x("UPDATE users SET role=? WHERE id=?", (new_role, user_id))
        self.log_action("change_role", f"Changed role for user '{username}' from '{old_role}' to '{new_role}'")

    def update_user_color(self, user_id: int, color: str):
        """Update a user's color preference."""
        self.x("UPDATE users SET fg_color=? WHERE id=?", (color, user_id))

    def delete_user(self, user_id: int):
        """Delete a user."""
        user = self.get_user_by_id(user_id)
        username = user.username if user else f"id={user_id}"
        self.x("DELETE FROM users WHERE id=?", (user_id,))
        self.log_action("delete_user", f"Deleted user '{username}'")

    def count_admins(self) -> int:
        """Count the number of admin users."""
        rows = self.q("SELECT COUNT(*) as cnt FROM users WHERE role=?", (ROLE_ADMIN,))
        return rows[0]["cnt"]


# =============================================================================
# Theme / UI basics
# =============================================================================


def cp(n: int) -> int:
    return curses.color_pair(n) if curses.has_colors() else 0


def init_theme(stdscr, db: Optional['DB'] = None, user: Optional['User'] = None):
    curses.curs_set(0)
    stdscr.keypad(True)

    if curses.has_colors():
        curses.start_color()
        try:
            curses.use_default_colors()
        except curses.error:
            pass

        # Get foreground color - user preference takes priority over global config
        fg_color = curses.COLOR_GREEN
        color_name = "green"

        if user and user.fg_color:
            color_name = user.fg_color.lower()
        elif db:
            color_name = db.get_config("fg_color", "green").lower()

        color_map = {
            "green": curses.COLOR_GREEN,
            "cyan": curses.COLOR_CYAN,
            "yellow": curses.COLOR_YELLOW,
            "white": curses.COLOR_WHITE,
        }
        # Handle custom color codes (stored as numeric values)
        if color_name.isdigit():
            fg_color = int(color_name)
        else:
            fg_color = color_map.get(color_name, curses.COLOR_GREEN)

        curses.init_pair(CP_NORMAL, fg_color, curses.COLOR_BLACK)
        curses.init_pair(CP_HILITE, curses.COLOR_BLACK, fg_color)
        curses.init_pair(CP_BAR, curses.COLOR_BLACK, fg_color)
        curses.init_pair(CP_DIM, fg_color, curses.COLOR_BLACK)

        stdscr.bkgd(" ", cp(CP_NORMAL))
        stdscr.attrset(cp(CP_NORMAL))
    else:
        stdscr.bkgd(" ", 0)
        stdscr.attrset(0)


def draw_chrome(stdscr, breadcrumb: str, footer: str, db_name: str = "", user: Optional['User'] = None, db: Optional['DB'] = None):
    # Get user and db_name from db if not explicitly passed
    if db is not None:
        if user is None and db.current_user is not None:
            user = db.current_user
        if not db_name and db.db_name:
            db_name = db.db_name

    stdscr.erase()
    h, w = stdscr.getmaxyx()

    # Check minimum window size
    MIN_HEIGHT = 10
    MIN_WIDTH = 40
    if h < MIN_HEIGHT or w < MIN_WIDTH:
        stdscr.clear()
        msg = f"Terminal too small! Need at least {MIN_WIDTH}x{MIN_HEIGHT}, have {w}x{h}"
        try:
            stdscr.addstr(0, 0, msg[:w-1] if w > 0 else "")
            if h > 1:
                stdscr.addstr(1, 0, "Please resize your terminal."[:w-1] if w > 0 else "")
        except curses.error:
            pass
        stdscr.refresh()
        time.sleep(0.1)
        return

    stdscr.attrset(cp(CP_BAR) | curses.A_BOLD)
    stdscr.addnstr(0, 0, " " * max(0, w - 1), max(0, w - 1))

    title = APP_TITLE
    if db_name.strip():
        title = f"{APP_TITLE} — {db_name.strip()}"

    head = f"{title} — {breadcrumb}"
    stdscr.addnstr(0, 1, head, max(0, w - 3))

    # Display user and version in top right corner
    if user:
        role_short = {"admin": "Admin", "editor": "Editor", "viewer": "Viewer"}.get(user.role, user.role)
        right_str = f"{user.username} ({role_short}) | v{VERSION} "
    else:
        right_str = f"v{VERSION} "
    right_x = max(0, w - len(right_str) - 1)
    stdscr.addnstr(0, right_x, right_str, len(right_str))

    stdscr.attrset(cp(CP_BAR))
    stdscr.addnstr(h - 1, 0, " " * max(0, w - 1), max(0, w - 1))
    stdscr.addnstr(h - 1, 1, footer, max(0, w - 3))
    stdscr.attrset(cp(CP_NORMAL))


def body_win(stdscr):
    h, w = stdscr.getmaxyx()
    return stdscr.derwin(h - 2, w, 1, 0)


def require_edit(stdscr, db: 'DB', breadcrumb: str, action: str = "edit") -> bool:
    """Check if current user can edit. Shows dialog and returns False if not."""
    user = db.current_user
    if user and user.is_viewer():
        dialog_message(stdscr, breadcrumb, "Read Only", [f"You do not have permission to {action}."], db_name=db.db_name)
        return False
    return True


def framed(win, title: str):
    win.erase()
    win.attrset(cp(CP_NORMAL))
    win.box()
    if title:
        t = f" {title} "
        maxy, maxx = win.getmaxyx()
        win.addnstr(0, 2, t, max(0, maxx - 4), cp(CP_BAR) | curses.A_BOLD)


def center_rect(stdscr, height: int, width: int) -> Tuple[int, int, int, int]:
    H, W = stdscr.getmaxyx()
    height = max(6, min(height, H - 2))
    width = max(30, min(width, W - 2))
    y = (H - height) // 2
    x = (W - width) // 2
    return y, x, height, width


def dialog_message(stdscr, breadcrumb: str, title: str, lines: List[str], db_name: str = ""):
    draw_chrome(stdscr, breadcrumb, "Enter/q: OK/back Esc: main menu", db_name=db_name)
    y, x, h, w = center_rect(
        stdscr,
        min(18, 7 + len(lines)),
        min(100, 10 + max((len(s) for s in lines), default=10)),
    )

    win = stdscr.derwin(h, w, y, x)
    framed(win, title)
    inner_w = w - 4
    for i, s in enumerate(lines[: h - 4]):
        win.addnstr(2 + i, 2, s, inner_w)

    win.refresh()
    while True:
        c = stdscr.getch()
        if c == 27:
            raise GoHome()
        if c in (ord("\n"), ord("\r"), curses.KEY_ENTER, ord("q"), ord("Q"), ord(" ")):
            break


def dialog_yes_no(stdscr, breadcrumb: str, title: str, lines: List[str], default_yes: bool = False, db_name: str = "") -> bool:
    draw_chrome(stdscr, breadcrumb, "y: Yes n: No Enter: default q: back (No) Esc: main menu", db_name=db_name)
    y, x, h, w = center_rect(
        stdscr,
        min(18, 8 + len(lines)),
        min(100, 12 + max((len(s) for s in lines), default=10)),
    )

    win = stdscr.derwin(h, w, y, x)
    framed(win, title)
    inner_w = w - 4
    for i, s in enumerate(lines[: h - 5]):
        win.addnstr(2 + i, 2, s, inner_w)
    win.addnstr(h - 3, 2, f"Default: {'Yes' if default_yes else 'No'}", inner_w, cp(CP_DIM) | curses.A_DIM)

    win.refresh()
    while True:
        c = stdscr.getch()
        if c == 27:
            raise GoHome()
        if c in (ord("q"), ord("Q")):
            return False  # 'q' means back/cancel, which is safer as False
        if c in (ord("y"), ord("Y")):
            return True
        if c in (ord("n"), ord("N")):
            return False
        if c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
            return default_yes


def dialog_danger_confirm(stdscr, breadcrumb: str, title: str, lines: List[str], db_name: str = "") -> bool:
    """A flashing danger confirmation dialog for destructive operations.
    User must type 'YES' (all caps) to confirm. Returns True only if confirmed."""
    draw_chrome(stdscr, breadcrumb, "Type YES to confirm, or q/Esc to cancel", db_name=db_name)
    y, x, h, w = center_rect(
        stdscr,
        min(20, 10 + len(lines)),
        min(100, 14 + max((len(s) for s in lines), default=10)),
    )

    win = stdscr.derwin(h, w, y, x)
    inner_w = w - 4

    input_buf = ""
    flash_state = True
    last_flash = time.time()

    # Use nodelay for non-blocking input during flash animation
    stdscr.nodelay(True)

    try:
        while True:
            # Update flash state every 0.5 seconds
            now = time.time()
            if now - last_flash >= 0.5:
                flash_state = not flash_state
                last_flash = now

            # Draw the dialog with flashing title
            win.erase()
            win.box()

            # Flashing title bar
            title_attr = (cp(CP_HILITE) | curses.A_BOLD | curses.A_REVERSE) if flash_state else (cp(CP_BAR) | curses.A_BOLD)
            t = f" ⚠ {title} ⚠ "
            maxy, maxx = win.getmaxyx()
            win.addnstr(0, 2, t, max(0, maxx - 4), title_attr)

            # Warning lines
            for i, s in enumerate(lines[: h - 7]):
                line_attr = cp(CP_NORMAL)
                if "WARNING" in s.upper() or "PERMANENT" in s.upper() or "CANNOT" in s.upper():
                    line_attr = cp(CP_NORMAL) | curses.A_BOLD
                win.addnstr(2 + i, 2, s, inner_w, line_attr)

            # Input prompt
            prompt_y = h - 4
            win.addnstr(prompt_y, 2, "Type YES to confirm:", inner_w, cp(CP_NORMAL) | curses.A_BOLD)

            # Input field with current text
            input_y = h - 3
            win.addnstr(input_y, 2, "> ", inner_w, cp(CP_NORMAL))
            input_attr = cp(CP_HILITE)
            win.addnstr(input_y, 4, input_buf + "_", inner_w - 4, input_attr)

            win.refresh()

            # Non-blocking input check
            c = stdscr.getch()

            if c == -1:
                # No input, just continue flashing
                time.sleep(0.05)
                continue

            if c == 27:  # Escape
                return False
            if c in (ord("q"), ord("Q")) and len(input_buf) == 0:
                return False
            if c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
                if input_buf == "YES":
                    return True
                else:
                    # Wrong input, flash and clear
                    input_buf = ""
            elif c in (curses.KEY_BACKSPACE, 127, 8):
                if input_buf:
                    input_buf = input_buf[:-1]
            elif 32 <= c <= 126:
                if len(input_buf) < 10:  # Limit input length
                    input_buf += chr(c)
    finally:
        # Restore blocking mode
        stdscr.nodelay(False)


def edit_line_dialog(
    stdscr, breadcrumb: str, title: str, label: str, initial: str = "", maxlen: int = 200, db_name: str = ""
) -> Optional[str]:
    draw_chrome(stdscr, breadcrumb, "Enter: OK q: back Esc: main menu", db_name=db_name)
    width = min(100, max(44, len(label) + 16))
    y, x, h, w = center_rect(stdscr, 10, width)
    win = stdscr.derwin(h, w, y, x)
    framed(win, title)
    win.addnstr(2, 2, label, w - 4)

    buf = list((initial or "")[:maxlen])
    pos = len(buf)
    while True:
        curses.curs_set(1)
        win.attrset(cp(CP_NORMAL))

        win.addnstr(4, 2, " " * (w - 4), w - 4)
        win.addnstr(4, 2, "> ", w - 4)

        text = "".join(buf)
        show = text[-(w - 7) :]
        win.attrset(cp(CP_HILITE))
        win.addnstr(4, 4, " " * (w - 6), w - 6)
        win.addnstr(4, 4, show, w - 6)

        win.attrset(cp(CP_NORMAL))
        win.move(4, min(w - 3, 4 + min(pos, w - 7)))
        win.refresh()

        c = stdscr.getch()
        if c == 27:
            curses.curs_set(0)
            raise GoHome()
        if c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
            curses.curs_set(0)
            return "".join(buf).strip()

        if c == curses.KEY_LEFT:
            pos = max(0, pos - 1)
        elif c == curses.KEY_RIGHT:
            pos = min(len(buf), pos + 1)
        elif c in (curses.KEY_BACKSPACE, 127, 8):
            if pos > 0:
                buf.pop(pos - 1)
                pos -= 1
        elif c == curses.KEY_DC:
            if pos < len(buf):
                buf.pop(pos)
        else:
            if 32 <= c <= 126 and len(buf) < maxlen:
                buf.insert(pos, chr(c))
                pos += 1


def password_dialog(
    stdscr, title: str, label: str, maxlen: int = 100, hint: Optional[List[str]] = None
) -> Optional[str]:
    """Password input dialog - shows asterisks instead of characters."""
    # Initialize basic colors for login screen
    if curses.has_colors():
        curses.start_color()
        try:
            curses.use_default_colors()
        except curses.error:
            pass
        curses.init_pair(CP_NORMAL, curses.COLOR_GREEN, curses.COLOR_BLACK)
        curses.init_pair(CP_HILITE, curses.COLOR_BLACK, curses.COLOR_GREEN)
        curses.init_pair(CP_BAR, curses.COLOR_BLACK, curses.COLOR_GREEN)
        curses.init_pair(CP_DIM, curses.COLOR_GREEN, curses.COLOR_BLACK)

    curses.curs_set(0)
    stdscr.clear()
    stdscr.bkgd(" ", cp(CP_NORMAL))

    hint_lines = hint if hint else []
    width = max(44, len(label) + 16)
    for line in hint_lines:
        width = max(width, len(line) + 6)
    width = min(90, width)
    height = 10 + len(hint_lines)

    y, x, h, w = center_rect(stdscr, height, width)
    win = stdscr.derwin(h, w, y, x)
    framed(win, title)
    win.addnstr(2, 2, label, w - 4)

    # Show hint lines at bottom of dialog
    for i, line in enumerate(hint_lines):
        row = h - 3 - len(hint_lines) + i
        if 0 < row < h - 1:
            win.addnstr(row, 2, line[:w-4], w - 4, cp(CP_DIM) | curses.A_DIM)

    buf = []
    input_row = 4
    while True:
        curses.curs_set(1)
        win.attrset(cp(CP_NORMAL))

        win.addnstr(input_row, 2, " " * (w - 4), w - 4)
        win.addnstr(input_row, 2, "> ", w - 4)

        # Show asterisks instead of actual characters
        show = "*" * min(len(buf), w - 7)
        win.attrset(cp(CP_HILITE))
        win.addnstr(input_row, 4, " " * (w - 6), w - 6)
        win.addnstr(input_row, 4, show, w - 6)

        win.attrset(cp(CP_NORMAL))
        win.move(input_row, min(w - 3, 4 + len(buf)))
        win.refresh()

        c = stdscr.getch()
        if c == 27:  # Escape
            curses.curs_set(0)
            return None
        if c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
            curses.curs_set(0)
            return "".join(buf)

        if c in (curses.KEY_BACKSPACE, 127, 8):
            if buf:
                buf.pop()
        elif 32 <= c <= 126 and len(buf) < maxlen:
            buf.append(chr(c))


def login_screen(stdscr, db: DB) -> Optional[User]:
    """Display login screen and authenticate user. Returns User or None if cancelled."""
    # Initialize basic colors for login screen
    if curses.has_colors():
        curses.start_color()
        try:
            curses.use_default_colors()
        except curses.error:
            pass
        curses.init_pair(CP_NORMAL, curses.COLOR_GREEN, curses.COLOR_BLACK)
        curses.init_pair(CP_HILITE, curses.COLOR_BLACK, curses.COLOR_GREEN)
        curses.init_pair(CP_BAR, curses.COLOR_BLACK, curses.COLOR_GREEN)
        curses.init_pair(CP_DIM, curses.COLOR_GREEN, curses.COLOR_BLACK)

    error_msg = ""

    while True:
        curses.curs_set(0)
        stdscr.clear()
        stdscr.bkgd(" ", cp(CP_NORMAL))

        h, w = stdscr.getmaxyx()

        # Draw title
        title = f"╔══ {APP_TITLE} v{VERSION} ══╗"
        title_x = (w - len(title)) // 2
        stdscr.addnstr(h // 2 - 6, max(0, title_x), title, w - 1, cp(CP_NORMAL) | curses.A_BOLD)

        # Draw login box
        box_w = 50
        box_h = 12
        box_y = h // 2 - 3
        box_x = (w - box_w) // 2

        try:
            win = stdscr.derwin(box_h, box_w, box_y, box_x)
        except curses.error:
            win = stdscr.derwin(min(box_h, h - box_y), min(box_w, w - box_x), box_y, max(0, box_x))

        framed(win, "Login")
        inner_w = box_w - 4

        win.addnstr(2, 2, "Username:", inner_w, cp(CP_NORMAL))
        win.addnstr(5, 2, "Password:", inner_w, cp(CP_NORMAL))

        if error_msg:
            win.addnstr(8, 2, error_msg, inner_w, cp(CP_NORMAL) | curses.A_BOLD)

        win.addnstr(10, 2, "Esc to quit", inner_w, cp(CP_DIM))

        win.refresh()
        stdscr.refresh()

        # Get username
        curses.curs_set(1)
        username_buf = []
        username_y = 3

        while True:
            win.addnstr(username_y, 2, " " * (inner_w), inner_w)
            win.addnstr(username_y, 2, "> ", inner_w)
            win.attrset(cp(CP_HILITE))
            win.addnstr(username_y, 4, " " * (inner_w - 4), inner_w - 4)
            win.addnstr(username_y, 4, "".join(username_buf), inner_w - 4)
            win.attrset(cp(CP_NORMAL))
            win.move(username_y, 4 + len(username_buf))
            win.refresh()

            c = stdscr.getch()
            if c == 27:  # Escape
                curses.curs_set(0)
                return None
            if c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
                break
            if c in (curses.KEY_BACKSPACE, 127, 8):
                if username_buf:
                    username_buf.pop()
            elif 32 <= c <= 126 and len(username_buf) < 50:
                username_buf.append(chr(c))

        username = "".join(username_buf).strip()
        if not username:
            error_msg = "Username required"
            continue

        # Get password
        password_buf = []
        password_y = 6

        while True:
            win.addnstr(password_y, 2, " " * (inner_w), inner_w)
            win.addnstr(password_y, 2, "> ", inner_w)
            win.attrset(cp(CP_HILITE))
            win.addnstr(password_y, 4, " " * (inner_w - 4), inner_w - 4)
            win.addnstr(password_y, 4, "*" * len(password_buf), inner_w - 4)
            win.attrset(cp(CP_NORMAL))
            win.move(password_y, 4 + len(password_buf))
            win.refresh()

            c = stdscr.getch()
            if c == 27:  # Escape
                curses.curs_set(0)
                return None
            if c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
                break
            if c in (curses.KEY_BACKSPACE, 127, 8):
                if password_buf:
                    password_buf.pop()
            elif 32 <= c <= 126 and len(password_buf) < 100:
                password_buf.append(chr(c))

        password = "".join(password_buf)
        curses.curs_set(0)

        # Authenticate
        user = db.authenticate(username, password)
        if user:
            db.set_current_user(user)
            db.log_action("login", f"User '{username}' logged in")
            return user
        else:
            error_msg = "Invalid username or password"


# =============================================================================
# Full-screen list & form screens
# =============================================================================


def render_list_rows(win, title: str, rows: List[ListRow], sel: int, top: int, ip_mode: bool = False):
    framed(win, title)
    h, w = win.getmaxyx()
    inner_w = w - 4
    view_h = h - 4

    # Check if any row has customer or location data
    has_customer = any(r.customer.strip() for r in rows)
    has_location = any(r.location.strip() for r in rows)
    show_columns = has_customer or has_location

    if show_columns:
        if ip_mode:
            # IP addresses: fixed narrow label (max 18 for "xxx.xxx.xxx.xxx"), more space for customer
            label_w = 18
            remaining = inner_w - label_w - 4
            cust_w = max(12, remaining * 2 // 3)
            loc_w = max(10, remaining - cust_w)
        else:
            # Default: balanced columns
            cust_w = min(24, max(10, inner_w // 5))
            loc_w = min(24, max(10, inner_w // 5))
            label_w = max(10, inner_w - cust_w - loc_w - 4)
        header = f"{'Item':<{label_w}} {'Customer':<{cust_w}} {'Location':<{loc_w}}"
    else:
        label_w = inner_w
        header = ""

    if header:
        win.addnstr(2, 2, header[:inner_w], inner_w, cp(CP_DIM) | curses.A_BOLD)

    for i in range(view_h - 1):
        idx = top + i
        y = 3 + i
        win.addnstr(y, 2, " " * inner_w, inner_w, cp(CP_NORMAL))
        if idx >= len(rows):
            continue
        r = rows[idx]
        if show_columns:
            s = f"{r.label[:label_w]:<{label_w}} {r.customer[:cust_w]:<{cust_w}} {r.location[:loc_w]:<{loc_w}}"
        else:
            s = r.label[:label_w]
        attr = cp(CP_HILITE) if idx == sel else cp(CP_NORMAL)
        win.addnstr(y, 2, s[:inner_w], inner_w, attr)


def full_screen_list(
    stdscr,
    breadcrumb: str,
    title: str,
    rows: List[ListRow],
    db_name: str = "",
    footer: str = "Up/Down: move PgUp/PgDn: page Enter: select q: back Esc: main menu",
    start_sel: int = 0,
    db: Optional['DB'] = None,
) -> Tuple[Optional[int], int]:
    draw_chrome(stdscr, breadcrumb, footer, db_name=db_name, db=db)
    bw = body_win(stdscr)
    sel = max(0, min(start_sel, max(0, len(rows) - 1)))
    top = 0
    while True:
        h, w = bw.getmaxyx()
        view_h = (h - 4) - 1
        if view_h < 1:
            view_h = 1
        if sel < top:
            top = sel
        if sel >= top + view_h:
            top = sel - view_h + 1

        render_list_rows(bw, title, rows, sel, top)
        bw.refresh()

        c = stdscr.getch()
        if c == 27:
            raise GoHome()
        if c in (ord("q"), ord("Q")):
            return None, sel

        if c in (curses.KEY_UP, ord("k")):
            sel = max(0, sel - 1)
        elif c in (curses.KEY_DOWN, ord("j")):
            sel = min(len(rows) - 1, sel + 1) if rows else 0
        elif c == curses.KEY_PPAGE:
            sel = max(0, sel - view_h)
        elif c == curses.KEY_NPAGE:
            sel = min(len(rows) - 1, sel + view_h) if rows else 0
        elif c == curses.KEY_HOME:
            sel = 0
        elif c == curses.KEY_END:
            sel = max(0, len(rows) - 1)
        elif c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
            return sel, sel


def horizontal_rule(win, y: int):
    h, w = win.getmaxyx()
    inner_w = w - 4
    if 0 <= y < h - 1:
        win.addnstr(y, 2, "-" * inner_w, inner_w, cp(CP_DIM) | curses.A_DIM)


def draw_attrs_block(win, title: str, attrs: Dict[str, str]):
    framed(win, title)
    h, w = win.getmaxyx()
    inner_w = w - 4

    norm = {k: (v or "") for k, v in attrs.items()}
    for k in STANDARD_KEYS:
        norm.setdefault(k, "")

    keys = STANDARD_KEYS + sorted([k for k in norm.keys() if k not in STANDARD_KEYS])
    y = 2
    for i, k in enumerate(keys):
        if y >= h - 2:
            break
        v = (norm.get(k, "") or "").strip()
        left = k
        right = v[:inner_w]
        rx = 2 + max(0, inner_w - len(right))
        win.addnstr(y, 2, " " * inner_w, inner_w, cp(CP_NORMAL))
        win.addnstr(y, 2, left[:inner_w], inner_w, cp(CP_NORMAL) | curses.A_BOLD)
        # Ensure we have positive width for the right side
        right_width = max(1, inner_w - (rx - 2))
        if rx + right_width <= w - 2:  # Make sure we don't exceed window bounds
            win.addnstr(y, rx, right, right_width, cp(CP_NORMAL))
        y += 1
        if i != len(keys) - 1 and y < h - 2:
            horizontal_rule(win, y)
            y += 1

    win.refresh()


def full_screen_form(
    stdscr,
    breadcrumb: str,
    title: str,
    fields: List[str],
    values: Dict[str, str],
    db_name: str = "",
    footer: str = "Up/Down: move Enter: edit d: delete key s: save q: back Esc: main menu",
) -> Optional[Dict[str, str]]:
    sel = 0
    while True:
        draw_chrome(stdscr, breadcrumb, footer, db_name=db_name)
        bw = body_win(stdscr)
        framed(bw, title)

        h, w = bw.getmaxyx()
        inner_w = w - 4
        view_h = h - 4

        key_w = min(20, max(10, inner_w // 4))
        val_w = max(10, inner_w - key_w - 3)

        for i in range(min(view_h, len(fields))):
            k = fields[i]
            v = values.get(k, "") or ""
            y = 2 + i
            attr = cp(CP_HILITE) if i == sel else cp(CP_NORMAL)
            bw.addnstr(y, 2, " " * inner_w, inner_w, attr)
            bw.addnstr(y, 2, f"{k:<{key_w}}", key_w, attr | curses.A_BOLD)
            bw.addnstr(y, 2 + key_w + 2, v[:val_w], val_w, attr)

        for y in range(2 + len(fields), h - 2):
            bw.addnstr(y, 2, " " * inner_w, inner_w, cp(CP_NORMAL))

        bw.refresh()
        c = stdscr.getch()
        if c == 27:
            raise GoHome()
        if c in (ord("q"), ord("Q")):
            return None

        if c in (curses.KEY_UP, ord("k")):
            sel = max(0, sel - 1)
        elif c in (curses.KEY_DOWN, ord("j")):
            sel = min(len(fields) - 1, sel + 1)
        elif c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
            k = fields[sel]
            newv = edit_line_dialog(stdscr, breadcrumb, "Edit field", f"{k}:", values.get(k, ""), db_name=db_name)
            if newv is None:
                continue
            values[k] = newv
        elif c in (ord("d"), ord("D")):
            k = fields[sel]
            if k not in STANDARD_KEYS and k not in VLAN_SUBNET_KEYS:
                confirm = dialog_yes_no(
                    stdscr,
                    breadcrumb,
                    "Delete key",
                    [f"Delete custom key '{k}'?"],
                    default_yes=False,
                    db_name=db_name,
                )
                if confirm:
                    fields = [f for f in fields if f != k]
                    if k in values:
                        values[k] = ""  # Mark for deletion by setting to empty
                    if sel >= len(fields) and len(fields) > 0:
                        sel = len(fields) - 1
        elif c in (ord("a"), ord("A")):
            k = edit_line_dialog(stdscr, breadcrumb, "Add key", "Key name:", "", db_name=db_name)
            if k:
                k = k.strip()
                if k and k not in STANDARD_KEYS and k not in VLAN_SUBNET_KEYS and k not in fields:
                    v = edit_line_dialog(stdscr, breadcrumb, "Add key", f"{k}:", "", db_name=db_name)
                    if v is not None:
                        fields.append(k)
                        values[k] = v
                        sel = len(fields) - 1
        elif c in (ord("s"), ord("S")):
            return values


# =============================================================================
# Selection helpers
# =============================================================================


def choose_vlan(stdscr, db: DB, breadcrumb: str, prompt_title: str, db_name: str) -> Optional[sqlite3.Row]:
    s = edit_line_dialog(stdscr, breadcrumb, prompt_title, "Enter VLAN (or Enter to list):", "", db_name=db_name)
    if s is None:
        return None

    if s.strip() == "":
        vlans = db.list_vlans()
        if not vlans:
            dialog_message(stdscr, breadcrumb, "No VLANs", ["No VLANs exist yet."], db_name=db_name)
            return None

        # Batch load aggregates for all VLANs
        vlan_ids = [v["id"] for v in vlans]
        aggregates = db.batch_aggregate_for_vlans(vlan_ids)

        rows = []
        for v in vlans:
            cust, loc = aggregates.get(v["id"], ("", ""))
            rows.append(ListRow(label=f"VLAN {v['vlan_num']} {v['name']}", customer=cust, location=loc))

        idx, _ = full_screen_list(stdscr, breadcrumb, "Select VLAN", rows, db_name=db_name)
        if idx is None:
            return None
        return vlans[idx]

    try:
        vlan_num = int(s.strip())
    except ValueError:
        dialog_message(stdscr, breadcrumb, "Error", ["VLAN must be an integer."], db_name=db_name)
        return None

    vlan = db.get_vlan_by_num(vlan_num)
    if not vlan:
        dialog_message(stdscr, breadcrumb, "Not found", [f"VLAN {vlan_num} not found."], db_name=db_name)
        return None
    return vlan


def choose_subnet(
    stdscr, db: DB, breadcrumb: str, title: str, db_name: str, only_with_ranges: bool = False
) -> Optional[sqlite3.Row]:
    s = edit_line_dialog(stdscr, breadcrumb, title, "CIDR prefix (or Enter to list):", "", db_name=db_name)
    if s is None:
        return None

    subnets = db.list_subnets_with_ranges() if only_with_ranges else db.list_subnets(None)

    # Pre-load all subnet ranges in one query for efficiency
    all_ranges: Dict[int, List[str]] = {}
    if subnets:
        for sub in subnets:
            all_ranges[sub["id"]] = []
        range_rows = db.q("SELECT bd_id, cidr FROM bd_ranges WHERE bd_id IN ({})".format(
            ','.join(['?'] * len(subnets))), tuple(sub["id"] for sub in subnets))
        for r in range_rows:
            all_ranges[r["bd_id"]].append(r["cidr"])

    if s.strip() != "":
        # Filter subnets by CIDR begins-with match
        search_term = s.strip()
        matching_subnets = []
        for sub in subnets:
            ranges = all_ranges.get(sub["id"], [])
            for cidr in ranges:
                if cidr.startswith(search_term):
                    matching_subnets.append(sub)
                    break
        subnets = matching_subnets

        if not subnets:
            dialog_message(stdscr, breadcrumb, "Not found", [f"No subnets with CIDR starting with '{s.strip()}'."], db_name=db_name)
            return None

    if not subnets:
        dialog_message(stdscr, breadcrumb, "No subnets", ["No subnets match this view."], db_name=db_name)
        return None

    # Batch load aggregates for all subnets
    bd_ids = [sub["id"] for sub in subnets]
    aggregates = db.batch_aggregate_for_subnets(bd_ids)

    rows: List[ListRow] = []
    for sub in subnets:
        sc, sl = aggregates.get(sub["id"], ("", ""))
        ranges = all_ranges.get(sub["id"], [])
        cidr_txt = ranges[0] if ranges else ""
        label = f"VLAN {sub['vlan_num']} {sub['name']}" + (f" {cidr_txt}" if cidr_txt else "")
        rows.append(ListRow(label=label, customer=sc, location=sl))

    idx, _ = full_screen_list(stdscr, breadcrumb, title, rows, db_name=db_name)
    if idx is None:
        return None
    return subnets[idx]


# =============================================================================
# Attribute editing logic
# =============================================================================


def set_attrs_with_overwrite_prompt(
    stdscr,
    db: DB,
    scope: str,
    scope_id: int,
    inheritable: int,
    new_values: Dict[str, str],
    breadcrumb: str,
    title: str,
    db_name: str,
):
    existing = db.get_attrs(scope, scope_id)
    existing = {k: (v or "") for k, v in existing.items()}
    for k, v in new_values.items():
        k = k.strip()
        if k == "":
            continue
        v = (v or "").strip()
        old = (existing.get(k, "") or "").strip()
        if old != "" and old != v:
            ok = dialog_yes_no(
                stdscr,
                breadcrumb,
                title,
                [f"{k} already set to:", old, "", "Overwrite with:", v],
                default_yes=False,
                db_name=db_name,
            )
            if not ok:
                continue

        if v == "":
            db.delete_attr(scope, scope_id, k)
        else:
            db.upsert_attr(scope, scope_id, k, v, inheritable)


def assign_subnet_inheritable_attrs_with_conflict_handling(
    stdscr,
    db: DB,
    bd_id: int,
    new_values: Dict[str, str],
    breadcrumb: str,
    db_name: str,
):
    set_attrs_with_overwrite_prompt(
        stdscr,
        db,
        scope="bd",
        scope_id=bd_id,
        inheritable=1,
        new_values=new_values,
        breadcrumb=breadcrumb,
        title="Save subnet attributes",
        db_name=db_name,
    )

    for k in STANDARD_KEYS:
        v = (new_values.get(k, "") or "").strip()
        if v == "":
            continue

        conflicts = db.q(
            """
            SELECT ip.id AS ip_id, ip.addr, a.value AS current_value
            FROM ip_addresses ip
            JOIN attributes a ON a.scope='ip' AND a.scope_id=ip.id AND a.key=?
            WHERE ip.bd_id=? AND a.value <> ?
            ORDER BY ip.addr
            """,
            (k, bd_id, v),
        )

        if not conflicts:
            continue

        ok = dialog_yes_no(
            stdscr,
            breadcrumb,
            "Override conflicts",
            [
                f"{len(conflicts)} IP(s) have conflicting '{k}' overrides.",
                "Overwrite those IP overrides to match the subnet value?",
            ],
            default_yes=False,
            db_name=db_name,
        )
        if not ok:
            continue

        for row in conflicts:
            db.upsert_attr("ip", row["ip_id"], k, v, 0)


def ensure_ip_linked(db: DB, ip_str: str) -> int:
    ip_id = db.ensure_ip(ip_str)
    res = db.resolve_for_ip(ip_str)
    db.set_ip_links(ip_id, res.vlan_id, res.bd_id)
    return ip_id


# =============================================================================
# "In use" vs "unused" helpers
# =============================================================================


def subnet_used_ips(db: DB, bd_id: int) -> List[str]:
    ip_rows = db.list_ip_rows_in_subnet(bd_id)
    if not ip_rows:
        return []

    # Build inherited attrs once (subnet + VLAN), same for every IP in this subnet
    subnet = db.get_subnet(bd_id)
    inherited = {}
    if subnet:
        inherited.update(db.get_attrs("vlan", subnet["vlan_id"]))
    inherited.update(db.get_attrs("bd", bd_id))
    for k in STANDARD_KEYS:
        inherited.setdefault(k, "")
    inherited = {k: (v or "").strip() for k, v in inherited.items()}

    # Batch-fetch all IP attributes in one query
    ip_id_map = {r["id"]: r["addr"] for r in ip_rows}
    ip_id_list = list(ip_id_map.keys())
    placeholders = ','.join(['?'] * len(ip_id_list))
    attr_rows = db.q(
        f"SELECT scope_id, key, value FROM attributes WHERE scope='ip' AND scope_id IN ({placeholders})",
        tuple(ip_id_list)
    )

    # Group attrs by IP id
    ip_attrs: Dict[int, Dict[str, str]] = {}
    for r in attr_rows:
        ip_attrs.setdefault(r["scope_id"], {})[r["key"]] = r["value"]

    # Compare each IP's attrs against inherited
    out = []
    for ip_id, ip_str in ip_id_map.items():
        attrs = ip_attrs.get(ip_id, {})
        nonempty = {k: v for k, v in attrs.items() if (v or "").strip()}
        if not nonempty:
            continue
        for k, v in nonempty.items():
            if inherited.get(k, "") != (v or "").strip():
                out.append(ip_str)
                break

    out.sort(key=lambda s: ipaddress.ip_address(s))
    return out


def subnet_unused_ips(db: DB, bd_id: int, used_set: set) -> Optional[List[str]]:
    ranges = [ipaddress.ip_network(r["cidr"], strict=False) for r in db.list_subnet_ranges(bd_id)]
    if not ranges:
        return []

    total = 0
    for n in ranges:
        try:
            total += sum(1 for _ in n.hosts())
        except TypeError:
            total += max(0, n.num_addresses - 2)

    if total > MAX_ENUM_HOSTS:
        return None

    out = set()
    for n in ranges:
        for ip in n.hosts():
            s = str(ip)
            if s not in used_set:
                out.add(s)

    return sorted(out, key=lambda s: ipaddress.ip_address(s))


# =============================================================================
# Workflows
# =============================================================================


def workflow_create_vlan(stdscr, db: DB, db_name: str):
    bc = "Main > Create VLAN"
    s = edit_line_dialog(stdscr, bc, "Create VLAN", "Enter VLAN:", "", db_name=db_name)
    if s is None or s.strip() == "":
        return

    try:
        vlan_num = int(s.strip())
    except ValueError:
        dialog_message(stdscr, bc, "Error", ["VLAN must be an integer."], db_name=db_name)
        return

    if db.vlan_exists(vlan_num):
        dialog_message(stdscr, bc, "Already exists", [f"VLAN {vlan_num} already exists."], db_name=db_name)
        return

    routed = dialog_yes_no(stdscr, bc, "Routed VLAN", ["Is this a routed VLAN?", "(Routed VLANs cannot have overlapping subnets with other routed VLANs)"], default_yes=False, db_name=db_name)
    routed_int = 1 if routed else 0

    values = {"name": "", "uplink": "", "customer": "", "location": "", "comment": ""}
    fields = ["name", "uplink"] + VLAN_SUBNET_KEYS
    out = full_screen_form(stdscr, bc, f"VLAN {vlan_num} info", fields, values, db_name=db_name)
    if out is None:
        return

    try:
        with db.transaction():
            vlan_id = db.create_vlan(vlan_num, out.get("name", ""), routed_int, out.get("uplink", ""))
            attrs = {k: out.get(k, "") for k in VLAN_SUBNET_KEYS}
            set_attrs_with_overwrite_prompt(stdscr, db, "vlan", vlan_id, 1, attrs, bc, "Save VLAN attributes", db_name=db_name)
    except ValueError as e:
        dialog_message(stdscr, bc, "Error", [str(e)], db_name=db_name)
        return
    except sqlite3.IntegrityError:
        dialog_message(stdscr, bc, "Error", [f"VLAN {vlan_num} already exists."], db_name=db_name)
        return

    dialog_message(stdscr, bc, "Created", [f"VLAN {vlan_num} created."], db_name=db_name)


def workflow_create_subnet(stdscr, db: DB, db_name: str):
    bc = "Main > Create Subnet"
    vlan = choose_vlan(stdscr, db, bc, "Create subnet", db_name=db_name)
    if not vlan:
        return

    name = edit_line_dialog(stdscr, bc, "Create Subnet", "Subnet name (optional):", "", db_name=db_name)
    if name is None:
        return

    cidr = edit_line_dialog(stdscr, bc, "Create Subnet", "CIDR:", "", db_name=db_name)
    if cidr is None or cidr.strip() == "":
        return

    try:
        bd_id = db.create_subnet(vlan["id"], name.strip())
    except sqlite3.IntegrityError:
        dialog_message(stdscr, bc, "Error", ["A subnet with that name already exists in this VLAN."], db_name=db_name)
        return

    try:
        added = db.add_subnet_range(bd_id, cidr.strip())
        if not added:
            dialog_message(stdscr, bc, "Duplicate", ["This CIDR already exists for this subnet."], db_name=db_name)
    except ValueError as e:
        dialog_message(stdscr, bc, "Error", [f"Invalid CIDR: {e}"], db_name=db_name)
        db.delete_subnet(bd_id)
        return

    values = {k: "" for k in VLAN_SUBNET_KEYS}
    out = full_screen_form(stdscr, bc, "Subnet attributes", VLAN_SUBNET_KEYS, values, db_name=db_name)
    if out is not None:
        assign_subnet_inheritable_attrs_with_conflict_handling(stdscr, db, bd_id, out, bc, db_name=db_name)

    dialog_message(stdscr, bc, "Created", [f"Subnet '{name.strip()}' created in VLAN {vlan['vlan_num']}."], db_name=db_name)


def workflow_edit_vlan(stdscr, db: DB, db_name: str):
    bc = "Main > Edit VLAN"
    vlan = choose_vlan(stdscr, db, bc, "Edit VLAN", db_name=db_name)
    if not vlan:
        return
    screen_vlan_menu(stdscr, db, vlan["id"], db_name=db_name)


def workflow_edit_subnet(stdscr, db: DB, db_name: str):
    bc = "Main > Edit Subnet"
    subnet = choose_subnet(stdscr, db, bc, "Select subnet", db_name=db_name, only_with_ranges=False)
    if not subnet:
        return
    screen_subnet_menu(stdscr, db, subnet["id"], db_name=db_name)


def workflow_edit_ip(stdscr, db: DB, db_name: str):
    bc = "Main > Edit IP"

    # IP-first: prompt immediately for an IP address.
    s = edit_line_dialog(
        stdscr,
        bc,
        "Edit IP",
        "IP address (blank to pick from a subnet):",
        "",
        db_name=db_name,
    )
    if s is None:
        return

    def subnet_picker_list_flow():
        """Subnet selector (CIDR required) + existing IP list picker flow."""

        subnet = choose_subnet(
            stdscr,
            db,
            bc,
            "Select subnet (CIDR required)",
            db_name=db_name,
            only_with_ranges=True,
        )
        if not subnet:
            return

        bd_id = subnet["id"]
        breadcrumb = f"Subnet {subnet['name']} > Edit IP"

        ranges = [ipaddress.ip_network(r["cidr"], strict=False) for r in db.list_subnet_ranges(bd_id)]
        if not ranges:
            dialog_message(stdscr, breadcrumb, "No CIDR", ["This subnet has no CIDR ranges configured."], db_name=db_name)
            return

        used = subnet_used_ips(db, bd_id)
        used_set = set(used)
        unused = subnet_unused_ips(db, bd_id, used_set)

        if unused is None:
            ip_rows = db.list_ip_rows_in_subnet(bd_id)
            candidates = sorted([r["addr"] for r in ip_rows], key=lambda x: ipaddress.ip_address(x))
            if not candidates:
                dialog_message(
                    stdscr,
                    breadcrumb,
                    "Too many addresses",
                    [
                        "Subnet too large to enumerate.",
                        "No stored IP objects to list.",
                        f"Limit is {MAX_ENUM_HOSTS} hosts.",
                    ],
                    db_name=db_name,
                )
                return
        else:
            candidates = sorted(set(used + unused), key=lambda x: ipaddress.ip_address(x))

        rows: List[ListRow] = []
        for ip_str in candidates:
            _, eff = db.effective_attrs_for_ip(ip_str)
            rows.append(ListRow(label=ip_str, customer=eff.get("Customer", ""), location=eff.get("Location", "")))

        pick, _ = full_screen_list(stdscr, breadcrumb, "Select IP", rows, db_name=db_name)
        if pick is None:
            return

        ip_str = rows[pick].label
        screen_edit_ip_in_subnet(stdscr, db, ip_str, breadcrumb, db_name=db_name)

    # Blank => existing subnet-based picker.
    if s.strip() == "":
        subnet_picker_list_flow()
        return

    # Search for IPs starting with the input
    search_str = s.strip()
    all_ips = db.q("SELECT addr FROM ip_addresses ORDER BY addr")
    matching_ips = [row["addr"] for row in all_ips if row["addr"].startswith(search_str)]

    if not matching_ips:
        # No matches, fall back to subnet picker
        subnet_picker_list_flow()
        return

    if len(matching_ips) == 1:
        # Single match, edit it directly
        ip_str = matching_ips[0]
    else:
        # Multiple matches, let user choose
        ip_rows = [ListRow(ip, "", "") for ip in matching_ips]
        idx, _ = full_screen_list(stdscr, bc, f"Select IP (matched '{search_str}')", ip_rows, db_name=db_name)
        if idx is None:
            return
        ip_str = matching_ips[idx]

    # Resolve the IP into a configured subnet
    res = db.resolve_for_ip(ip_str)
    if res.bd_id is None:
        subnet_picker_list_flow()
        return

    breadcrumb = f"Subnet {res.bd_name} > Edit IP" if res.bd_name else bc
    screen_edit_ip_in_subnet(stdscr, db, ip_str, breadcrumb, db_name=db_name)

def workflow_configure(stdscr, db: DB, user: User, db_name: str) -> str:
    bc = "Main > Configure"

    while True:
        # Check if BIOS settings are unlocked
        bios_unlocked = db.get_config('bios_unlocked', '0') == '1'

        options = [
            ListRow("My Color Preference", "", ""),
            ListRow("Change My Password", "", ""),
        ]

        # Admin-only options
        if user.is_admin():
            options.append(ListRow("User Management", "", ""))
            options.append(ListRow("Database Name", "", ""))
            options.append(ListRow("Snapshot Settings", "", ""))

        # Add BIOS settings if unlocked
        if bios_unlocked:
            options.append(ListRow("BIOS Boot Settings", "", ""))

        options.append(ListRow("Back", "", ""))
        back_idx = len(options) - 1

        idx, _ = full_screen_list(stdscr, bc, "Configure", options, db_name=db_name)
        if idx is None or idx == back_idx:
            return db_name

        if idx == 0:
            # My Color Preference
            current_color = user.fg_color or "green"
            color_options = [
                ListRow("Green (default)", "", ""),
                ListRow("Cyan", "", ""),
                ListRow("Yellow", "", ""),
                ListRow("White", "", ""),
            ]

            color_idx, _ = full_screen_list(stdscr, bc, f"Select Color (current: {current_color})", color_options, db_name=db_name)
            if color_idx is not None:
                color_names = ["green", "cyan", "yellow", "white"]
                color_name = color_names[color_idx]
                db.update_user_color(user.id, color_name)
                # Update the user object
                user = db.get_user_by_id(user.id)
                db.set_current_user(user)
                init_theme(stdscr, db, user)
                dialog_message(stdscr, bc, "Color Changed", [f"Your color preference set to {color_name}"], db_name=db_name)

        elif idx == 1:
            # Change My Password
            workflow_change_password(stdscr, db, user, db_name)

        elif idx == 2 and user.is_admin():
            # User Management (admin only)
            workflow_user_management(stdscr, db, db_name)

        elif idx == 3 and user.is_admin():
            # Database name (admin only)
            s = edit_line_dialog(
                stdscr,
                bc,
                "Configure",
                "Database name (shown in header):",
                initial=db.get_config("db_name", ""),
                db_name=db_name,
            )
            if s is not None:
                db.set_config("db_name", s.strip())
                db_name = db.get_config("db_name", "").strip()
                db.db_name = db_name

        elif idx == 4 and user.is_admin():
            # Snapshot settings (admin only)
            workflow_snapshot_settings(stdscr, db, db_name)

        elif bios_unlocked and ((user.is_admin() and idx == 5) or (not user.is_admin() and idx == 2)):
            # BIOS settings (only available if unlocked)
            workflow_bios_settings(stdscr, db, db_name)

    return db_name


def workflow_change_password(stdscr, db: DB, user: User, db_name: str):
    """Change current user's password."""
    bc = "Main > Configure > Change Password"

    # Get current password
    current_pw = password_dialog(stdscr, "Change Password", "Current password:")
    if current_pw is None:
        return

    # Verify current password
    if db.authenticate(user.username, current_pw) is None:
        dialog_message(stdscr, bc, "Error", ["Current password is incorrect."], db_name=db_name)
        return

    # Get new password with hint
    new_pw = password_dialog(stdscr, "Change Password", "New password:", hint=db.PASSWORD_REQUIREMENTS)
    if new_pw is None or not new_pw:
        return

    # Validate new password
    valid, error = db.validate_password(new_pw)
    if not valid:
        dialog_message(stdscr, bc, "Invalid Password", [error], db_name=db_name)
        return

    # Confirm new password
    confirm_pw = password_dialog(stdscr, "Change Password", "Confirm new password:")
    if confirm_pw is None:
        return

    if new_pw != confirm_pw:
        dialog_message(stdscr, bc, "Error", ["Passwords do not match."], db_name=db_name)
        return

    # Update password
    db.update_user_password(user.id, new_pw)
    dialog_message(stdscr, bc, "Success", ["Password changed successfully."], db_name=db_name)


def workflow_user_management(stdscr, db: DB, db_name: str):
    """Manage users (admin only)."""
    bc = "Main > Configure > Users"

    while True:
        users = db.list_users()

        rows = []
        for u in users:
            role_display = {"admin": "Admin", "editor": "Editor", "viewer": "Viewer"}.get(u.role, u.role)
            rows.append(ListRow(f"{u.username} ({role_display})", "", ""))
        rows.append(ListRow("+ Add New User", "", ""))
        rows.append(ListRow("Back", "", ""))

        idx, _ = full_screen_list(stdscr, bc, "User Management", rows, db_name=db_name,
                                   footer="Enter: edit user a: add d: delete q: back")

        if idx is None or idx == len(rows) - 1:
            return

        if idx == len(rows) - 2:
            # Add new user
            workflow_add_user(stdscr, db, db_name)
            continue

        # Edit existing user
        selected_user = users[idx]
        workflow_edit_user(stdscr, db, selected_user, db_name)


def workflow_add_user(stdscr, db: DB, db_name: str):
    """Add a new user."""
    bc = "Main > Configure > Users > Add"

    # Get username
    username = edit_line_dialog(stdscr, bc, "Add User", "Username:", "", db_name=db_name)
    if username is None or not username.strip():
        return
    username = username.strip()

    # Validate username
    valid, error = db.validate_username(username)
    if not valid:
        dialog_message(stdscr, bc, "Invalid Username", [error], db_name=db_name)
        return

    # Check if username exists
    if db.get_user_by_username(username):
        dialog_message(stdscr, bc, "Error", [f"Username '{username}' already exists."], db_name=db_name)
        return

    # Get password with hint
    password = password_dialog(stdscr, "Add User", "Password:", hint=db.PASSWORD_REQUIREMENTS)
    if password is None or not password:
        dialog_message(stdscr, bc, "Error", ["Password is required."], db_name=db_name)
        return

    # Validate password
    valid, error = db.validate_password(password)
    if not valid:
        dialog_message(stdscr, bc, "Invalid Password", [error], db_name=db_name)
        return

    # Confirm password
    confirm_password = password_dialog(stdscr, "Add User", "Confirm password:")
    if confirm_password is None:
        return

    if password != confirm_password:
        dialog_message(stdscr, bc, "Error", ["Passwords do not match."], db_name=db_name)
        return

    # Select role
    role_options = [
        ListRow("Admin - Full access including user management", "", ""),
        ListRow("Editor - Full access except user management", "", ""),
        ListRow("Viewer - Read-only access", "", ""),
    ]
    role_idx, _ = full_screen_list(stdscr, bc, "Select Role", role_options, db_name=db_name)
    if role_idx is None:
        return

    roles = [ROLE_ADMIN, ROLE_EDITOR, ROLE_VIEWER]
    role = roles[role_idx]

    # Create user
    db.create_user(username, password, role)
    dialog_message(stdscr, bc, "Success", [f"User '{username}' created."], db_name=db_name)


def workflow_edit_user(stdscr, db: DB, target_user: User, db_name: str):
    """Edit an existing user."""
    bc = f"Main > Configure > Users > {target_user.username}"

    while True:
        role_display = {"admin": "Admin", "editor": "Editor", "viewer": "Viewer"}.get(target_user.role, target_user.role)

        options = [
            ListRow(f"Change Role (current: {role_display})", "", ""),
            ListRow("Reset Password", "", ""),
            ListRow("Delete User", "", ""),
            ListRow("Back", "", ""),
        ]

        idx, _ = full_screen_list(stdscr, bc, f"Edit User: {target_user.username}", options, db_name=db_name)
        if idx is None or idx == 3:
            return

        if idx == 0:
            # Change role
            role_options = [
                ListRow("Admin - Full access including user management", "", ""),
                ListRow("Editor - Full access except user management", "", ""),
                ListRow("Viewer - Read-only access", "", ""),
            ]
            role_idx, _ = full_screen_list(stdscr, bc, "Select New Role", role_options, db_name=db_name)
            if role_idx is not None:
                roles = [ROLE_ADMIN, ROLE_EDITOR, ROLE_VIEWER]
                new_role = roles[role_idx]

                # Prevent removing the last admin
                if target_user.role == ROLE_ADMIN and new_role != ROLE_ADMIN:
                    if db.count_admins() <= 1:
                        dialog_message(stdscr, bc, "Error", ["Cannot remove the last admin.", "Create another admin first."], db_name=db_name)
                        continue

                db.update_user_role(target_user.id, new_role)
                target_user = db.get_user_by_id(target_user.id)
                dialog_message(stdscr, bc, "Success", [f"Role changed to {new_role}."], db_name=db_name)

        elif idx == 1:
            # Reset password
            new_pw = password_dialog(stdscr, "Reset Password", "New password:", hint=db.PASSWORD_REQUIREMENTS)
            if new_pw and new_pw.strip():
                # Validate new password
                valid, error = db.validate_password(new_pw)
                if not valid:
                    dialog_message(stdscr, bc, "Invalid Password", [error], db_name=db_name)
                    continue
                db.update_user_password(target_user.id, new_pw)
                dialog_message(stdscr, bc, "Success", ["Password reset successfully."], db_name=db_name)

        elif idx == 2:
            # Delete user
            # Prevent deleting yourself
            if db.current_user and target_user.id == db.current_user.id:
                dialog_message(stdscr, bc, "Error", ["Cannot delete your own account."], db_name=db_name)
                continue
            # Prevent deleting the last admin
            if target_user.role == ROLE_ADMIN and db.count_admins() <= 1:
                dialog_message(stdscr, bc, "Error", ["Cannot delete the last admin."], db_name=db_name)
                continue

            confirm = dialog_yes_no(
                stdscr, bc, "Delete User",
                [f"Delete user '{target_user.username}'?", "This cannot be undone."],
                default_yes=False, db_name=db_name
            )
            if confirm:
                db.delete_user(target_user.id)
                dialog_message(stdscr, bc, "Deleted", [f"User '{target_user.username}' deleted."], db_name=db_name)
                return


def workflow_snapshot_settings(stdscr, db: DB, db_name: str):
    """Manage snapshot settings."""
    bc = "Main > Configure > Snapshots"

    while True:
        # Get current settings
        max_count = db.get_config("snapshot_max_count", "20")
        enabled = db.get_config("snapshot_enabled", "1") == "1"
        enabled_str = "Enabled" if enabled else "Disabled"

        options = [
            ListRow(f"Max Snapshots (current: {max_count})", "", ""),
            ListRow(f"Auto-snapshot on changes: {enabled_str}", "", ""),
            ListRow("View Snapshot Statistics", "", ""),
            ListRow("Clean Up Old Snapshots Now", "", ""),
            ListRow("Back", "", ""),
        ]

        idx, _ = full_screen_list(stdscr, bc, "Snapshot Settings", options, db_name=db_name)
        if idx is None or idx == 4:
            return

        if idx == 0:
            # Max snapshots
            new_max = edit_line_dialog(
                stdscr, bc, "Max Snapshots",
                "Maximum number of snapshots to keep:",
                initial=max_count,
                db_name=db_name
            )
            if new_max and new_max.strip().isdigit():
                db.set_config("snapshot_max_count", new_max.strip())
                dialog_message(stdscr, bc, "Updated", [f"Max snapshots set to {new_max.strip()}"], db_name=db_name)

        elif idx == 1:
            # Toggle auto-snapshot
            new_enabled = not enabled
            db.set_config("snapshot_enabled", "1" if new_enabled else "0")
            status = "enabled" if new_enabled else "disabled"
            dialog_message(stdscr, bc, "Updated", [f"Auto-snapshot {status}"], db_name=db_name)

        elif idx == 2:
            # View statistics
            stats = db.get_snapshot_stats()
            lines = [
                f"Total snapshots: {stats['count']}",
                f"Total size: {stats['total_size_mb']:.2f} MB",
                f"Oldest: {stats['oldest_date'] or 'N/A'}",
                f"Newest: {stats['newest_date'] or 'N/A'}",
            ]
            dialog_message(stdscr, bc, "Snapshot Statistics", lines, db_name=db_name)

        elif idx == 3:
            # Manual cleanup
            confirm = dialog_yes_no(
                stdscr, bc, "Clean Up Snapshots",
                ["This will remove old snapshots according to the retention policy.", "", "Continue?"],
                default_yes=False,
                db_name=db_name
            )
            if confirm:
                before_count = db.q("SELECT COUNT(*) as cnt FROM snapshots")[0]["cnt"]
                db.cleanup_old_snapshots()
                after_count = db.q("SELECT COUNT(*) as cnt FROM snapshots")[0]["cnt"]
                deleted = before_count - after_count
                dialog_message(
                    stdscr, bc, "Cleanup Complete",
                    [f"Deleted {deleted} old snapshot(s)", f"Remaining: {after_count}"],
                    db_name=db_name
                )


def workflow_search(stdscr, db: DB, db_name: str):
    bc = "Main > Search"
    options = [
        ListRow("VLAN", "", ""),
        ListRow("IP address", "", ""),
        ListRow("CIDR", "", ""),
        ListRow("Customer", "", ""),
        ListRow("Location", "", ""),
    ]

    idx, _ = full_screen_list(stdscr, bc, "Search", options, db_name=db_name)
    if idx is None:
        return

    if idx == 0:
        s = edit_line_dialog(stdscr, bc, "Search VLAN", "VLAN:", "", db_name=db_name)
        if s is None or s.strip() == "":
            return
        try:
            vlan_num = int(s.strip())
        except ValueError:
            dialog_message(stdscr, bc, "Error", ["VLAN must be an integer."], db_name=db_name)
            return

        vlan = db.get_vlan_by_num(vlan_num)
        if not vlan:
            dialog_message(stdscr, bc, "Not found", [f"VLAN {vlan_num} not found."], db_name=db_name)
            return

        screen_vlan_menu(stdscr, db, vlan["id"], db_name=db_name)
        return

    if idx == 1:
        ipstr = edit_line_dialog(stdscr, bc, "Search IP", "IP address:", "", db_name=db_name)
        if ipstr is None or ipstr.strip() == "":
            return
        try:
            ipaddress.ip_address(ipstr.strip())
        except ValueError as e:
            dialog_message(stdscr, bc, "Error", [f"Invalid IP: {e}"], db_name=db_name)
            return

        res = db.resolve_for_ip(ipstr.strip())
        if res.bd_id is None:
            dialog_message(stdscr, bc, "Not found", ["No matching subnet found for that IP."], db_name=db_name)
            return

        # Check if IP is in the used list to set the correct view
        normalized_ip = str(ipaddress.ip_address(ipstr.strip()))
        used_ips = subnet_used_ips(db, res.bd_id)
        show_in_use = normalized_ip in used_ips

        screen_subnet_menu(
            stdscr,
            db,
            res.bd_id,
            highlight_ip=normalized_ip,
            show_in_use_default=show_in_use,
            db_name=db_name,
        )
        return

    if idx == 2:
        cidr = edit_line_dialog(stdscr, bc, "Search CIDR", "CIDR:", "", db_name=db_name)
        if cidr is None or cidr.strip() == "":
            return
        try:
            net = ipaddress.ip_network(cidr.strip(), strict=False)
        except ValueError as e:
            dialog_message(stdscr, bc, "Error", [f"Invalid CIDR: {e}"], db_name=db_name)
            return

        candidates = []
        for r in db.q(
            """
            SELECT br.id AS br_id, br.cidr, bd.id AS bd_id, bd.name, bd.vlan_id,
                   v.vlan_num AS vlan_num, v.name AS vlan_name
            FROM bd_ranges br
            JOIN broadcast_domains bd ON bd.id=br.bd_id
            JOIN vlans v ON v.id=bd.vlan_id
            """
        ):
            try:
                parent = ipaddress.ip_network(r["cidr"], strict=False)
            except Exception:
                continue

            if net.subnet_of(parent) or parent.subnet_of(net) or net == parent:
                candidates.append(r)

        if not candidates:
            dialog_message(stdscr, bc, "Not found", ["No matching subnet found for that CIDR."], db_name=db_name)
            return

        if len(candidates) == 1:
            screen_subnet_menu(stdscr, db, candidates[0]["bd_id"], show_in_use_default=True, db_name=db_name)
            return

        rows = []
        for c in candidates:
            bd_id = c["bd_id"]
            sc, sl = db.batch_aggregate_for_subnets([bd_id]).get(bd_id, ("", ""))
            label = f"VLAN {c['vlan_num']} {c['name']} {c['cidr']}"
            rows.append(ListRow(label=label, customer=sc, location=sl))

        pick, _ = full_screen_list(stdscr, bc, "Select subnet", rows, db_name=db_name)
        if pick is None:
            return

        screen_subnet_menu(stdscr, db, candidates[pick]["bd_id"], show_in_use_default=True, db_name=db_name)
        return

    key = "customer" if idx == 3 else "location"
    moderows = [ListRow("regex", "", ""), ListRow("starts with", "", ""), ListRow("contains", "", "")]
    mode, _ = full_screen_list(stdscr, bc, f"Match mode ({key})", moderows, db_name=db_name)
    if mode is None:
        return

    pat = edit_line_dialog(stdscr, bc, f"Search {key}", "Pattern:", "", db_name=db_name)
    if pat is None or pat == "":
        return

    def match(s: str) -> bool:
        s = s or ""
        if mode == 0:
            try:
                return re.search(pat, s) is not None
            except re.error as e:
                raise ValueError(f"Invalid regex pattern: {e}")
        if mode == 1:
            return s.startswith(pat)
        return pat in s

    try:
        # Batch load all VLAN aggregates
        vlans = db.list_vlans()
        vlan_ids = [v["id"] for v in vlans]
        vlan_aggregates = db.batch_aggregate_for_vlans(vlan_ids)

        vlanhits = []
        for v in vlans:
            cust, loc = vlan_aggregates.get(v["id"], ("", ""))
            if key == "customer" and match(cust):
                vlanhits.append(("vlan", v["id"], f"VLAN {v['vlan_num']} {v['name']}", cust, loc))
            if key == "location" and match(loc):
                vlanhits.append(("vlan", v["id"], f"VLAN {v['vlan_num']} {v['name']}", cust, loc))

        # Batch load all subnet aggregates
        subnets = db.list_subnets(None)
        bd_ids = [s["id"] for s in subnets]
        subnet_aggregates = db.batch_aggregate_for_subnets(bd_ids)

        subnethits = []
        for s in subnets:
            cust, loc = subnet_aggregates.get(s["id"], ("", ""))
            label = f"VLAN {s['vlan_num']} {s['name']}"
            if key == "customer" and match(cust):
                subnethits.append(("bd", s["id"], label, cust, loc))
            if key == "location" and match(loc):
                subnethits.append(("bd", s["id"], label, cust, loc))
    except ValueError as e:
        dialog_message(stdscr, bc, "Error", [str(e)], db_name=db_name)
        return

    hits = vlanhits + subnethits
    if not hits:
        dialog_message(stdscr, bc, "No results", ["No matches."], db_name=db_name)
        return

    rows = [ListRow(label=h[2], customer=h[3], location=h[4]) for h in hits]
    pick, _ = full_screen_list(stdscr, bc, "Results", rows, db_name=db_name)
    if pick is None:
        return

    scope, scope_id = hits[pick][0], hits[pick][1]
    if scope == "vlan":
        screen_vlan_menu(stdscr, db, scope_id, db_name=db_name)
    else:
        screen_subnet_menu(stdscr, db, scope_id, db_name=db_name)


def workflow_list(stdscr, db: DB, db_name: str):
    bc = "Main > List"
    choices = [ListRow("VLANs", "", ""), ListRow("Subnets", "", ""), ListRow("IP addresses (stored)", "", "")]
    idx, _ = full_screen_list(stdscr, bc, "List", choices, db_name=db_name)
    if idx is None:
        return

    if idx == 0:
        vlans = db.list_vlans()
        if not vlans:
            dialog_message(stdscr, bc, "No VLANs", ["No VLANs exist yet."], db_name=db_name)
            return

        # Batch load aggregates
        vlan_ids = [v["id"] for v in vlans]
        aggregates = db.batch_aggregate_for_vlans(vlan_ids)

        rows = []
        for v in vlans:
            cust, loc = aggregates.get(v["id"], ("", ""))
            rows.append(ListRow(label=f"VLAN {v['vlan_num']} {v['name']}", customer=cust, location=loc))

        pick, _ = full_screen_list(stdscr, bc, "VLANs", rows, db_name=db_name)
        if pick is None:
            return
        screen_vlan_menu(stdscr, db, vlans[pick]["id"], db_name=db_name)
        return

    if idx == 1:
        subnets = db.list_subnets(None)
        if not subnets:
            dialog_message(stdscr, bc, "No subnets", ["No subnets exist yet."], db_name=db_name)
            return

        # Batch load aggregates and ranges
        bd_ids = [s["id"] for s in subnets]
        aggregates = db.batch_aggregate_for_subnets(bd_ids)

        # Pre-load all ranges in one query
        all_ranges: Dict[int, List[str]] = {bid: [] for bid in bd_ids}
        range_rows = db.q("SELECT bd_id, cidr FROM bd_ranges WHERE bd_id IN ({})".format(
            ','.join(['?'] * len(bd_ids))), tuple(bd_ids))
        for r in range_rows:
            all_ranges[r["bd_id"]].append(r["cidr"])

        rows = []
        for s in subnets:
            sc, sl = aggregates.get(s["id"], ("", ""))
            ranges = all_ranges.get(s["id"], [])
            cidr_txt = ranges[0] if ranges else ""
            label = f"VLAN {s['vlan_num']} {s['name']}" + (f" {cidr_txt}" if cidr_txt else "")
            rows.append(ListRow(label=label, customer=sc, location=sl))

        pick, _ = full_screen_list(stdscr, bc, "Subnets", rows, db_name=db_name)
        if pick is None:
            return

        screen_subnet_menu(stdscr, db, subnets[pick]["id"], db_name=db_name)
        return

    ips = db.q("SELECT ip.id, ip.addr, ip.bd_id, ip.vlan_id FROM ip_addresses ip ORDER BY ip.addr")
    if not ips:
        dialog_message(stdscr, bc, "No IPs", ["No IP objects exist yet."], db_name=db_name)
        return

    # Batch-fetch Customer/Location attrs for all IPs in one query
    ip_id_list = [r["id"] for r in ips]
    placeholders = ','.join(['?'] * len(ip_id_list))
    attr_rows = db.q(
        f"SELECT scope_id, key, value FROM attributes WHERE scope='ip' AND scope_id IN ({placeholders}) AND key IN ('Customer', 'Location')",
        tuple(ip_id_list)
    )
    ip_attrs: Dict[int, Dict[str, str]] = {}
    for r in attr_rows:
        ip_attrs.setdefault(r["scope_id"], {})[r["key"]] = r["value"] or ""

    # Build inherited attrs cache by (bd_id, vlan_id) pair
    inherited_cache: Dict[Tuple[Optional[int], Optional[int]], Tuple[str, str]] = {}
    for r in ips:
        cache_key = (r["bd_id"], r["vlan_id"])
        if cache_key not in inherited_cache:
            cust, loc = "", ""
            if r["bd_id"]:
                bd_attrs = db.get_attrs("bd", r["bd_id"])
                cust = bd_attrs.get("Customer", "")
                loc = bd_attrs.get("Location", "")
            if r["vlan_id"] and (not cust or not loc):
                vlan_attrs = db.get_attrs("vlan", r["vlan_id"])
                if not cust:
                    cust = vlan_attrs.get("Customer", "")
                if not loc:
                    loc = vlan_attrs.get("Location", "")
            inherited_cache[cache_key] = (cust, loc)

    rows = []
    ipaddrs = []
    for r in ips:
        iid = r["id"]
        ip_cust = ip_attrs.get(iid, {}).get("Customer", "")
        ip_loc = ip_attrs.get(iid, {}).get("Location", "")
        inh_cust, inh_loc = inherited_cache.get((r["bd_id"], r["vlan_id"]), ("", ""))
        cust = ip_cust or inh_cust
        loc = ip_loc or inh_loc
        rows.append(ListRow(label=r["addr"], customer=cust, location=loc))
        ipaddrs.append(r["addr"])

    pick, _ = full_screen_list(stdscr, bc, "IP addresses (stored)", rows, db_name=db_name)
    if pick is None:
        return

    ip = ipaddrs[pick]
    res = db.resolve_for_ip(ip)
    if res.bd_id is not None:
        screen_subnet_menu(stdscr, db, res.bd_id, highlight_ip=ip, show_in_use_default=True, db_name=db_name)
    else:
        screen_edit_ip_in_subnet(stdscr, db, ip, bc, db_name=db_name)


# =============================================================================
# Screen: VLAN menu
# =============================================================================


def _handle_vlan_menu_key(stdscr, db: DB, c: int, vlan_id: int, vlan: sqlite3.Row, breadcrumb: str, db_name: str, sort_by_subnet: bool, subnet_sel: int) -> Tuple[Optional[sqlite3.Row], bool, int, bool]:
    """Handle common key presses for VLAN menu. Returns (updated_vlan, sort_by_subnet, subnet_sel, was_deleted)."""
    was_deleted = False
    bail = vlan, sort_by_subnet, subnet_sel, was_deleted

    if c in (ord("a"), ord("A")):
        if not require_edit(stdscr, db, breadcrumb):
            return bail
        current = db.get_attrs("vlan", vlan_id)
        uplink_val = vlan["uplink"] if vlan["uplink"] else ""
        vals = {"name": vlan["name"], "uplink": uplink_val, **{k: current.get(k, "") for k in VLAN_SUBNET_KEYS}}
        out = full_screen_form(stdscr, breadcrumb, "Edit VLAN", ["name", "uplink"] + VLAN_SUBNET_KEYS, vals, db_name=db_name)
        if out is not None:
            db.x("UPDATE vlans SET name=?, uplink=? WHERE id=?", (out.get("name", ""), out.get("uplink", ""), vlan_id))
            assign = {k: out.get(k, "") for k in VLAN_SUBNET_KEYS}
            set_attrs_with_overwrite_prompt(
                stdscr, db, "vlan", vlan_id, 1, assign, breadcrumb, "Save VLAN attributes", db_name=db_name
            )
            vlan = db.get_vlan_by_id(vlan_id)

    elif c in (ord("r"), ord("R")):
        if not require_edit(stdscr, db, breadcrumb):
            return bail
        current_routed = vlan['routed']
        new_routed = not current_routed

        if new_routed:
            # Changing to routed, check for conflicts
            conflict = db.check_vlan_can_be_routed(vlan_id)
            if conflict:
                dialog_message(
                    stdscr,
                    breadcrumb,
                    "Cannot mark as routed",
                    [
                        f"Subnet {conflict[0]} overlaps with",
                        f"VLAN {conflict[1]} subnet {conflict[2]}",
                        "(both would be routed VLANs)"
                    ],
                    db_name=db_name
                )
                return bail

        db.update_vlan_routed(vlan_id, 1 if new_routed else 0)
        vlan = db.get_vlan_by_id(vlan_id)

    elif c in (ord("s"), ord("S")):
        sort_by_subnet = not sort_by_subnet
        subnet_sel = 0

    elif c in (ord("x"), ord("X")):
        workflow_export_vlan(stdscr, db, vlan_id, db_name=db_name)

    elif c in (ord("n"), ord("N")):
        if not require_edit(stdscr, db, breadcrumb):
            return bail
        name = edit_line_dialog(stdscr, breadcrumb, "Add subnet", "Subnet name (optional):", "", db_name=db_name)
        if name is not None:
            cidr = edit_line_dialog(stdscr, breadcrumb, "Add subnet", "CIDR:", "", db_name=db_name)
            if cidr and cidr.strip():
                try:
                    bd_id = db.create_subnet(vlan_id, name.strip())
                    try:
                        db.add_subnet_range(bd_id, cidr.strip())
                    except ValueError as e:
                        dialog_message(stdscr, breadcrumb, "Error", [f"Invalid CIDR: {e}"], db_name=db_name)
                        db.delete_subnet(bd_id)
                except sqlite3.IntegrityError:
                    dialog_message(stdscr, breadcrumb, "Error", ["Subnet name already exists in this VLAN."], db_name=db_name)

    elif c in (ord("d"), ord("D")):
        if not require_edit(stdscr, db, breadcrumb, "delete"):
            return bail
        # Delete VLAN with danger confirmation
        subnets = db.list_subnets(vlan_id)
        subnet_count = len(subnets)
        ip_count = sum(len(db.list_ip_rows_in_subnet(s["id"])) for s in subnets)

        warning_lines = [
            f"VLAN: {vlan['vlan_num']} - {vlan['name']}",
            "",
            f"This will permanently delete:",
            f"  • {subnet_count} subnet(s)",
            f"  • {ip_count} IP address(es)",
            f"  • All associated attributes",
            "",
            "WARNING: This action CANNOT be undone!",
            "(A snapshot will be created before deletion)",
        ]

        confirm = dialog_danger_confirm(
            stdscr,
            breadcrumb,
            "DELETE VLAN",
            warning_lines,
            db_name=db_name
        )

        if confirm:
            deleted_subnets, deleted_ips = db.delete_vlan(vlan_id)
            dialog_message(
                stdscr,
                breadcrumb,
                "VLAN Deleted",
                [
                    f"VLAN {vlan['vlan_num']} has been deleted.",
                    f"Removed {deleted_subnets} subnet(s) and {deleted_ips} IP(s).",
                ],
                db_name=db_name
            )
            was_deleted = True

    return vlan, sort_by_subnet, subnet_sel, was_deleted


def screen_vlan_menu(stdscr, db: DB, vlan_id: int, start_subnet_idx: int = 0, db_name: str = ""):
    vlan = db.get_vlan_by_id(vlan_id)
    if not vlan:
        return

    breadcrumb = f"VLAN {vlan['vlan_num']}"
    subnet_sel = start_subnet_idx
    sort_by_subnet = False  # False = sort by name, True = sort by subnet (CIDR)

    while True:
        draw_chrome(
            stdscr,
            breadcrumb,
            "Enter: open a: edit n: add subnet r: routed s: sort x: export d: DELETE q: back Esc: main",
            db_name=db_name,
        )

        bw = body_win(stdscr)
        H, W = bw.getmaxyx()
        bw.erase()

        left_w = max(40, int(W * 0.45))
        right_w = W - left_w

        left = bw.derwin(H, left_w, 0, 0)
        right = bw.derwin(H, right_w, 0, left_w)

        framed(left, "VLAN details")
        framed(right, "Subnets")

        inner_w = left_w - 4
        cust, loc = db.batch_aggregate_for_vlans([vlan_id]).get(vlan_id, ("", ""))
        routed_str = "Yes" if vlan['routed'] else "No"
        uplink_str = vlan["uplink"] if vlan["uplink"] else ""

        left.addnstr(2, 2, f"VLAN: {vlan['vlan_num']} {vlan['name']}", inner_w, cp(CP_NORMAL) | curses.A_BOLD)
        left.addnstr(3, 2, f"Routed: {routed_str}", inner_w)
        left.addnstr(4, 2, f"Uplink: {uplink_str}", inner_w)
        left.addnstr(5, 2, f"Customer: {cust}", inner_w)
        left.addnstr(6, 2, f"Location: {loc}", inner_w)

        attrs_win_h = max(8, H - 10)
        attrs_win = left.derwin(attrs_win_h, left_w - 2, 8, 1)
        draw_attrs_block(attrs_win, "Attributes", db.get_attrs("vlan", vlan_id))

        subnets = db.list_subnets(vlan_id)

        # Batch load aggregates and ranges for all subnets
        if subnets:
            bd_ids = [s["id"] for s in subnets]
            aggregates = db.batch_aggregate_for_subnets(bd_ids)

            # Pre-load all ranges in one query
            all_ranges: Dict[int, List[str]] = {bid: [] for bid in bd_ids}
            range_rows = db.q("SELECT bd_id, cidr FROM bd_ranges WHERE bd_id IN ({})".format(
                ','.join(['?'] * len(bd_ids))), tuple(bd_ids))
            for r in range_rows:
                all_ranges[r["bd_id"]].append(r["cidr"])
        else:
            aggregates = {}
            all_ranges = {}

        rows = []
        for s in subnets:
            sc, sl = aggregates.get(s["id"], ("", ""))
            ranges_s = all_ranges.get(s["id"], [])
            cidr_txt = ranges_s[0] if ranges_s else ""
            label = f"{s['name']}" + (f" {cidr_txt}" if cidr_txt else "")
            rows.append(ListRow(label=label, customer=sc, location=sl))

        # Sort based on current sort mode
        if sort_by_subnet:
            # Sort by CIDR (extract from label after space, or use full label if no space)
            subnets = sorted(zip(subnets, rows), key=lambda x: x[1].label.split()[-1] if ' ' in x[1].label else x[1].label)
        else:
            # Sort by name
            subnets = sorted(zip(subnets, rows), key=lambda x: x[0]['name'])

        # Unzip back into separate lists
        if subnets:
            subnets, rows = zip(*subnets)
            subnets = list(subnets)
            rows = list(rows)
        else:
            subnets = []
            rows = []

        if not rows:
            right.addnstr(2, 2, "No subnets.", right_w - 4, cp(CP_DIM) | curses.A_DIM)
            right.refresh()
            left.refresh()
            bw.refresh()

            c = stdscr.getch()
            if c == 27:
                raise GoHome()
            if c in (ord("q"), ord("Q")):
                return

            vlan, sort_by_subnet, subnet_sel, was_deleted = _handle_vlan_menu_key(
                stdscr, db, c, vlan_id, vlan, breadcrumb, db_name, sort_by_subnet, subnet_sel
            )
            if was_deleted:
                return
            continue

        sel = max(0, min(subnet_sel, len(rows) - 1))
        top = 0
        subnet_sel = sel
        render_list_rows(right, "Subnets", rows, sel, top)
        right.refresh()
        left.refresh()
        bw.refresh()

        c = stdscr.getch()
        if c == 27:
            raise GoHome()
        if c in (ord("q"), ord("Q")):
            return

        if c in (curses.KEY_UP, ord("k")):
            subnet_sel = max(0, subnet_sel - 1)
        elif c in (curses.KEY_DOWN, ord("j")):
            subnet_sel = min(len(rows) - 1, subnet_sel + 1)
        elif c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
            bd_id = subnets[subnet_sel]["id"]
            screen_subnet_menu(stdscr, db, bd_id, db_name=db_name)
        else:
            vlan, sort_by_subnet, subnet_sel, was_deleted = _handle_vlan_menu_key(
                stdscr, db, c, vlan_id, vlan, breadcrumb, db_name, sort_by_subnet, subnet_sel
            )
            if was_deleted:
                return


# =============================================================================
# Screen: Edit IP attributes
# =============================================================================


def screen_edit_ip_in_subnet(stdscr, db: DB, ip_str: str, breadcrumb: str, db_name: str):
    if not require_edit(stdscr, db, breadcrumb):
        return

    ip_id = ensure_ip_linked(db, ip_str)
    current = db.get_attrs("ip", ip_id)
    fields = STANDARD_KEYS + sorted([k for k in current.keys() if k not in STANDARD_KEYS])
    values = {k: current.get(k, "") for k in fields}

    out = full_screen_form(
        stdscr,
        breadcrumb,
        f"IP {ip_str} attributes",
        fields,
        values,
        db_name=db_name,
        footer="Up/Down: move Enter: edit s: save a: add key q: back Esc: main menu",
    )
    if out is None:
        return

    assign = {k: out.get(k, "") for k in STANDARD_KEYS}
    set_attrs_with_overwrite_prompt(stdscr, db, "ip", ip_id, 0, assign, breadcrumb, "Save IP attributes", db_name=db_name)

    extras = {k: out.get(k, "") for k in out.keys() if k not in STANDARD_KEYS}
    set_attrs_with_overwrite_prompt(stdscr, db, "ip", ip_id, 0, extras, breadcrumb, "Save IP attributes", db_name=db_name)


# =============================================================================
# Screen: Subnet menu
# =============================================================================


def _handle_subnet_command(stdscr, db: DB, c: int, bd_id: int, subnet, vlan, breadcrumb: str, db_name: str) -> Tuple[Any, Any, bool, bool]:
    """Handle e/n/m/d key commands for subnet menu. Returns (subnet, vlan, need_rebuild, was_deleted)."""
    need_rebuild = False

    if c in (ord("e"), ord("E")):
        if not require_edit(stdscr, db, breadcrumb):
            return subnet, vlan, False, False
        current = db.get_attrs("bd", bd_id)
        out = full_screen_form(
            stdscr, breadcrumb, "Edit subnet", VLAN_SUBNET_KEYS,
            {k: current.get(k, "") for k in VLAN_SUBNET_KEYS}, db_name=db_name,
        )
        if out is not None:
            assign_subnet_inheritable_attrs_with_conflict_handling(stdscr, db, bd_id, out, breadcrumb, db_name=db_name)
            subnet = db.get_subnet(bd_id)
            vlan = db.get_vlan_by_id(subnet["vlan_id"])
            need_rebuild = True

    elif c in (ord("n"), ord("N")):
        if not require_edit(stdscr, db, breadcrumb):
            return subnet, vlan, False, False
        new_name = edit_line_dialog(stdscr, breadcrumb, "Rename subnet", "New subnet name:", subnet['name'], db_name=db_name)
        if new_name is not None:
            try:
                db.update_subnet_name(bd_id, new_name.strip())
                subnet = db.get_subnet(bd_id)
            except sqlite3.IntegrityError:
                dialog_message(stdscr, breadcrumb, "Error", ["A subnet with that name already exists in this VLAN."], db_name=db_name)

    elif c in (ord("m"), ord("M")):
        if not require_edit(stdscr, db, breadcrumb):
            return subnet, vlan, False, False
        new_vlan = choose_vlan(stdscr, db, breadcrumb, "Move subnet to VLAN", db_name=db_name)
        if new_vlan:
            db.move_subnet_to_vlan(bd_id, new_vlan["id"])
            subnet = db.get_subnet(bd_id)
            vlan = db.get_vlan_by_id(subnet["vlan_id"])
            need_rebuild = True

    elif c in (ord("d"), ord("D")):
        if not require_edit(stdscr, db, breadcrumb, "delete"):
            return subnet, vlan, False, False
        confirm = dialog_yes_no(
            stdscr, breadcrumb, "Delete subnet",
            [f"Delete subnet '{subnet['name']}'?", "This will remove all IP addresses in this subnet."],
            default_yes=False, db_name=db_name,
        )
        if confirm:
            db.delete_subnet(bd_id)
            return subnet, vlan, False, True

    return subnet, vlan, need_rebuild, False


def screen_subnet_menu(
    stdscr,
    db: DB,
    bd_id: int,
    highlight_ip: Optional[str] = None,
    show_in_use_default: bool = True,
    db_name: str = "",
):
    subnet = db.get_subnet(bd_id)
    if not subnet:
        return

    vlan = db.get_vlan_by_id(subnet["vlan_id"])
    breadcrumb = f"Subnet {subnet['name']} (VLAN {vlan['vlan_num']})"

    show_in_use = show_in_use_default
    sel = 0
    top = 0

    # Cache for IP list and rows - only rebuild when needed
    cached_rows: Optional[List[ListRow]] = None
    cached_ip_list: Optional[List[str]] = None
    cache_mode: Optional[bool] = None  # Track which mode the cache is for
    need_rebuild = True  # Flag to trigger cache rebuild

    while True:
        # Rebuild cache only when needed
        if need_rebuild or cache_mode != show_in_use:
            used = subnet_used_ips(db, bd_id)
            used_set = set(used)

            if show_in_use:
                ip_list = used
            else:
                unused = subnet_unused_ips(db, bd_id, used_set)
                if unused is None:
                    dialog_message(
                        stdscr,
                        breadcrumb,
                        "Too many addresses",
                        [
                            "This subnet is too large to enumerate unused addresses safely.",
                            f"Limit is {MAX_ENUM_HOSTS} hosts.",
                            "Switching back to In use view.",
                        ],
                        db_name=db_name,
                    )
                    show_in_use = True
                    ip_list = used
                else:
                    ip_list = unused

            # Build rows with effective attributes
            rows: List[ListRow] = []

            # Batch load IP attributes for better performance
            if ip_list:
                # Get all IP rows that exist in database
                ip_ids = {}
                for ip in ip_list:
                    ip_row = db.get_ip_row(ip)
                    if ip_row:
                        ip_ids[ip] = ip_row["id"]

                # Batch load IP attributes
                if ip_ids:
                    ip_id_list = list(ip_ids.values())
                    placeholders = ','.join(['?'] * len(ip_id_list))
                    attr_rows = db.q(
                        f"SELECT scope_id, key, value FROM attributes WHERE scope='ip' AND scope_id IN ({placeholders}) AND key IN ('Customer', 'Location')",
                        tuple(ip_id_list)
                    )
                    ip_attrs: Dict[int, Dict[str, str]] = {iid: {} for iid in ip_id_list}
                    for r in attr_rows:
                        ip_attrs[r["scope_id"]][r["key"]] = r["value"] or ""
                else:
                    ip_attrs = {}

                # Get inherited attributes from subnet/vlan (same for all IPs in this subnet)
                inherited_cust = ""
                inherited_loc = ""
                bd_attrs = db.get_attrs("bd", bd_id)
                if bd_attrs.get("Customer"):
                    inherited_cust = bd_attrs["Customer"]
                if bd_attrs.get("Location"):
                    inherited_loc = bd_attrs["Location"]
                if not inherited_cust or not inherited_loc:
                    vlan_attrs = db.get_attrs("vlan", subnet["vlan_id"])
                    if not inherited_cust and vlan_attrs.get("Customer"):
                        inherited_cust = vlan_attrs["Customer"]
                    if not inherited_loc and vlan_attrs.get("Location"):
                        inherited_loc = vlan_attrs["Location"]

                for ip in ip_list:
                    if ip in ip_ids:
                        iid = ip_ids[ip]
                        cust = ip_attrs.get(iid, {}).get("Customer") or inherited_cust
                        loc = ip_attrs.get(iid, {}).get("Location") or inherited_loc
                    else:
                        cust = inherited_cust
                        loc = inherited_loc
                    rows.append(ListRow(label=ip, customer=cust, location=loc))

            cached_rows = rows
            cached_ip_list = ip_list
            cache_mode = show_in_use
            need_rebuild = False

            # Handle highlight_ip on first load
            if highlight_ip is not None and highlight_ip in ip_list:
                sel = ip_list.index(highlight_ip)
                highlight_ip = None

        rows = cached_rows
        ip_list = cached_ip_list

        draw_chrome(
            stdscr,
            breadcrumb,
            "t: toggle in-use/unused Enter: edit IP e: edit subnet n: rename subnet m: move VLAN d: delete q: back Esc: main menu",
            db_name=db_name,
        )

        bw = body_win(stdscr)
        H, W = bw.getmaxyx()
        bw.erase()

        left_w = max(44, int(W * 0.48))
        right_w = W - left_w

        left = bw.derwin(H, left_w, 0, 0)
        right = bw.derwin(H, right_w, 0, left_w)

        framed(left, "Subnet details")
        framed(right, "IP addresses")

        inner_w = left_w - 4
        ranges = [r["cidr"] for r in db.list_subnet_ranges(bd_id)]
        cust, loc = db.batch_aggregate_for_subnets([bd_id]).get(bd_id, ("", ""))

        left.addnstr(2, 2, f"Subnet: {subnet['name']}", inner_w, cp(CP_NORMAL) | curses.A_BOLD)
        left.addnstr(3, 2, f"VLAN: {vlan['vlan_num']} {vlan['name']}", inner_w)
        left.addnstr(4, 2, f"Customer: {cust}", inner_w)
        left.addnstr(5, 2, f"Location: {loc}", inner_w)

        left.addnstr(7, 2, "Ranges:", inner_w, cp(CP_DIM) | curses.A_BOLD)
        y = 8
        for cidr in ranges[: max(0, H - 14)]:
            left.addnstr(y, 4, cidr, inner_w - 2)
            y += 1

        # Calculate space for three sections: subnet attrs + selected IP attrs
        remaining_h = H - (y + 2)
        subnet_attrs_h = max(4, remaining_h // 2)
        ip_attrs_h = max(4, remaining_h - subnet_attrs_h - 1)

        # Subnet attributes section
        attrs_win = left.derwin(subnet_attrs_h, left_w - 2, y + 1, 1)
        draw_attrs_block(attrs_win, "Attributes", db.get_attrs("bd", bd_id))

        mode_txt = "In use" if show_in_use else "Unused"

        if not rows:
            right.addnstr(4, 2, "No addresses to display.", right_w - 4, cp(CP_DIM) | curses.A_DIM)
            right.refresh()
            left.refresh()
            bw.refresh()

            c = stdscr.getch()
            if c == 27:
                raise GoHome()
            if c in (ord("q"), ord("Q")):
                return

            if c in (ord("t"), ord("T")):
                show_in_use = not show_in_use
                sel = 0
                top = 0
                need_rebuild = True

            elif c in (ord("e"), ord("E"), ord("n"), ord("N"), ord("m"), ord("M"), ord("d"), ord("D")):
                subnet, vlan, need_rebuild, was_deleted = _handle_subnet_command(stdscr, db, c, bd_id, subnet, vlan, breadcrumb, db_name)
                if was_deleted:
                    return

            continue

        sel = max(0, min(sel, len(rows) - 1))

        # Draw selected IP attributes section
        selected_ip = rows[sel].label
        ip_attrs_y = y + 1 + subnet_attrs_h + 1
        ip_attrs_win = left.derwin(ip_attrs_h, left_w - 2, ip_attrs_y, 1)
        ip_row = db.get_ip_row(selected_ip)
        if ip_row:
            draw_attrs_block(ip_attrs_win, f"{selected_ip} Attributes", db.get_attrs("ip", ip_row["id"]))
        else:
            draw_attrs_block(ip_attrs_win, f"{selected_ip} Attributes", {})

        # Calculate scrolling
        H_right, _ = right.getmaxyx()
        view_h = (H_right - 4) - 1
        if view_h < 1:
            view_h = 1
        if sel < top:
            top = sel
        if sel >= top + view_h:
            top = sel - view_h + 1

        render_list_rows(right, f"IP addresses ({mode_txt})", rows, sel, top, ip_mode=True)
        right.refresh()
        left.refresh()
        bw.refresh()

        c = stdscr.getch()
        if c == 27:
            raise GoHome()
        if c in (ord("q"), ord("Q")):
            return

        if c in (ord("t"), ord("T")):
            show_in_use = not show_in_use
            sel = 0
            top = 0
            need_rebuild = True

        elif c in (curses.KEY_UP, ord("k")):
            sel = max(0, sel - 1)
        elif c in (curses.KEY_DOWN, ord("j")):
            sel = min(len(rows) - 1, sel + 1)
        elif c == curses.KEY_PPAGE:
            sel = max(0, sel - view_h)
        elif c == curses.KEY_NPAGE:
            sel = min(len(rows) - 1, sel + view_h)
        elif c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
            ip_str = rows[sel].label
            screen_edit_ip_in_subnet(stdscr, db, ip_str, breadcrumb, db_name=db_name)
            need_rebuild = True  # Rebuild after editing in case attributes changed

        elif c in (ord("e"), ord("E"), ord("n"), ord("N"), ord("m"), ord("M"), ord("d"), ord("D")):
            subnet, vlan, need_rebuild, was_deleted = _handle_subnet_command(stdscr, db, c, bd_id, subnet, vlan, breadcrumb, db_name)
            if was_deleted:
                return


# =============================================================================
# Export / Import
# =============================================================================


def export_vlans_to_xlsx(db: DB, vlan_ids: Optional[List[int]] = None, filename: str = "") -> str:
    """Export VLANs to XLSX format. Returns filename."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")

    if not filename:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"ipam_export_{timestamp}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Get all VLANs to export
    if vlan_ids is None:
        vlans = db.list_vlans()
    else:
        vlans = [db.get_vlan_by_id(vid) for vid in vlan_ids]
        vlans = [v for v in vlans if v]

    # Get all custom keys
    custom_keys = db.get_all_custom_keys()
    all_keys = STANDARD_KEYS + custom_keys

    for vlan in vlans:
        vlan_id = vlan["id"]
        sheet_name = f"VLAN_{vlan['vlan_num']}_{vlan['name']}"[:31]  # XLSX sheet name limit
        sheet = wb.create_sheet(sheet_name)

        # Row 1: VLAN metadata
        vlan_attrs = db.get_attrs("vlan", vlan_id)
        routed_str = "Yes" if vlan["routed"] else "No"
        uplink_str = vlan["uplink"] if vlan["uplink"] else ""
        sheet['A1'] = "VLAN Metadata"
        sheet['A1'].font = Font(bold=True)
        sheet['B1'] = f"VLAN {vlan['vlan_num']}"
        sheet['C1'] = f"Name: {vlan['name']}"
        sheet['D1'] = f"Routed: {routed_str}"
        sheet['E1'] = f"Uplink: {uplink_str}"
        col = 6
        for key in VLAN_SUBNET_KEYS:
            sheet.cell(1, col, f"{key}: {vlan_attrs.get(key, '')}")
            col += 1

        # Row 2: Blank

        # Row 3: Headers
        headers = ["Subnet Name", "CIDR", "IP Address"] + all_keys
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(3, col_idx, header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Row 4+: Data
        row = 4
        subnets = db.list_subnets(vlan_id)
        for subnet in subnets:
            subnet_id = subnet["id"]
            subnet_attrs = db.get_attrs("bd", subnet_id)
            ranges = [r["cidr"] for r in db.list_subnet_ranges(subnet_id)]
            range_nets = []
            for cidr in ranges:
                try:
                    range_nets.append((cidr, ipaddress.ip_network(cidr, strict=False)))
                except ValueError:
                    range_nets.append((cidr, None))

            # Get all IPs in this subnet
            ip_rows = db.list_ip_rows_in_subnet(subnet_id)

            if ip_rows:
                # Track which ranges have IPs so we can export empty ranges too
                ranges_with_ips = set()

                for ip_row in ip_rows:
                    ip_str = ip_row["addr"]
                    ip_id = ip_row["id"]
                    ip_attrs = db.get_attrs("ip", ip_id)

                    # Find the most specific matching CIDR range for this IP
                    ip_cidr = ranges[0] if ranges else ""
                    try:
                        ip_obj = ipaddress.ip_address(ip_str)
                        best_prefix = -1
                        for cidr, net in range_nets:
                            if net and ip_obj in net and net.prefixlen > best_prefix:
                                ip_cidr = cidr
                                best_prefix = net.prefixlen
                    except ValueError:
                        pass
                    ranges_with_ips.add(ip_cidr)

                    sheet.cell(row, 1, subnet["name"])
                    sheet.cell(row, 2, ip_cidr)
                    sheet.cell(row, 3, ip_str)

                    # Only write IP attributes if they're non-empty (not inherited)
                    for col_idx, key in enumerate(all_keys, 4):
                        val = ip_attrs.get(key, "").strip()
                        if val:  # Only write non-empty values
                            sheet.cell(row, col_idx, val)

                    row += 1

                # Export any ranges that had no IPs in them
                for cidr in ranges:
                    if cidr not in ranges_with_ips:
                        sheet.cell(row, 1, subnet["name"])
                        sheet.cell(row, 2, cidr)
                        for col_idx, key in enumerate(all_keys, 4):
                            val = subnet_attrs.get(key, "").strip()
                            if val:
                                sheet.cell(row, col_idx, f"[Subnet: {val}]")
                        row += 1
            else:
                # Empty subnet — write a row for each range
                if ranges:
                    for cidr in ranges:
                        sheet.cell(row, 1, subnet["name"])
                        sheet.cell(row, 2, cidr)
                        for col_idx, key in enumerate(all_keys, 4):
                            val = subnet_attrs.get(key, "").strip()
                            if val:
                                sheet.cell(row, col_idx, f"[Subnet: {val}]")
                        row += 1
                else:
                    sheet.cell(row, 1, subnet["name"])
                    row += 1

        # Auto-size columns using get_column_letter for proper column names (A, B, ..., Z, AA, AB, etc.)
        for col_idx in range(1, len(headers) + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = 15

    wb.save(filename)
    return filename


def import_from_xlsx(stdscr, db: DB, filename: str, mode: str, db_name: str) -> Tuple[int, int, List[str]]:
    """Import from XLSX file. Returns (added, updated, errors).
    Modes: 'interactive', 'prefer_db', 'prefer_import', 'overwrite'"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ImportError("openpyxl is required for XLSX import. Install with: pip install openpyxl")

    wb = load_workbook(filename, data_only=True)
    added = 0
    updated = 0
    errors = []

    # Create a snapshot before import so the entire operation can be rolled back
    db.log_action("import_start", f"Import from {filename}", create_snapshot=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Parse VLAN info from sheet name
        if not sheet_name.startswith("VLAN_"):
            errors.append(f"Skipping sheet '{sheet_name}' - doesn't start with VLAN_")
            continue

        parts = sheet_name.split("_", 2)
        if len(parts) < 2:
            errors.append(f"Skipping sheet '{sheet_name}' - invalid format")
            continue

        try:
            vlan_num = int(parts[1])
        except ValueError:
            errors.append(f"Skipping sheet '{sheet_name}' - invalid VLAN number")
            continue

        # Read VLAN metadata from row 1
        vlan_name = ""
        vlan_routed = 0
        vlan_uplink = ""
        vlan_attrs = {}

        for col_idx in range(2, 20):  # Check first 20 columns
            cell_val = sheet.cell(1, col_idx).value
            if cell_val and isinstance(cell_val, str):
                if cell_val.lower().startswith("name:"):
                    vlan_name = cell_val.split(":", 1)[1].strip()
                elif cell_val.lower().startswith("routed:"):
                    routed_str = cell_val.split(":", 1)[1].strip().lower()
                    vlan_routed = 1 if routed_str == "yes" else 0
                elif cell_val.lower().startswith("uplink:"):
                    vlan_uplink = cell_val.split(":", 1)[1].strip()
                elif ":" in cell_val:
                    key, val = cell_val.split(":", 1)
                    vlan_attrs[key.strip()] = val.strip()

        # Check if VLAN exists
        existing_vlan = db.get_vlan_by_num(vlan_num)

        if existing_vlan:
            vlan_id = existing_vlan["id"]

            # Handle routed flag change
            if existing_vlan["routed"] != vlan_routed:
                if vlan_routed == 1:
                    conflict = db.check_vlan_can_be_routed(vlan_id)
                    if conflict:
                        errors.append(f"VLAN {vlan_num}: Cannot mark as routed - subnet {conflict[0]} overlaps with VLAN {conflict[1]} subnet {conflict[2]}")
                        continue
                db.update_vlan_routed(vlan_id, vlan_routed)
                updated += 1

            # Handle uplink change
            existing_uplink = existing_vlan['uplink'] if 'uplink' in existing_vlan.keys() else ""
            if existing_uplink != vlan_uplink:
                db.update_vlan_uplink(vlan_id, vlan_uplink)
                updated += 1
        else:
            # Create new VLAN
            try:
                vlan_id = db.create_vlan(vlan_num, vlan_name, vlan_routed, vlan_uplink)
                added += 1
            except Exception as e:
                errors.append(f"VLAN {vlan_num}: Failed to create - {e}")
                continue

        # Update VLAN attributes
        for key, val in vlan_attrs.items():
            db.upsert_attr("vlan", vlan_id, key, val, 1)

        # Read headers from row 3
        headers = []
        for col_idx in range(1, sheet.max_column + 1):
            header = sheet.cell(3, col_idx).value
            if header:
                headers.append((col_idx, str(header)))

        # Find column indices
        subnet_col = next((c for c, h in headers if h.lower() == "subnet name"), None)
        cidr_col = next((c for c, h in headers if h.lower() == "cidr"), None)
        ip_col = next((c for c, h in headers if h.lower() == "ip address"), None)
        attr_cols = [(c, h) for c, h in headers if h in STANDARD_KEYS or h in db.get_all_custom_keys()]

        if not subnet_col or not cidr_col or not ip_col:
            errors.append(f"VLAN {vlan_num}: Missing required columns")
            continue

        # Process data rows
        for row_idx in range(4, sheet.max_row + 1):
            subnet_name = sheet.cell(row_idx, subnet_col).value
            cidr = sheet.cell(row_idx, cidr_col).value
            ip_addr = sheet.cell(row_idx, ip_col).value

            if not subnet_name and not cidr and not ip_addr:
                continue  # Empty row

            subnet_name = str(subnet_name or "").strip()
            cidr = str(cidr or "").strip()
            ip_addr = str(ip_addr or "").strip()

            if not cidr:
                errors.append(f"VLAN {vlan_num} row {row_idx}: Missing CIDR")
                continue

            # Find or create subnet
            subnets = db.list_subnets(vlan_id)
            subnet = next((s for s in subnets if s["name"] == subnet_name), None)

            if not subnet:
                try:
                    bd_id = db.create_subnet(vlan_id, subnet_name)
                    added += 1
                except Exception as e:
                    errors.append(f"VLAN {vlan_num} row {row_idx}: Failed to create subnet - {e}")
                    continue
            else:
                bd_id = subnet["id"]

            # Add CIDR to subnet if not exists
            try:
                existing_ranges = [r["cidr"] for r in db.list_subnet_ranges(bd_id)]
                if cidr not in existing_ranges:
                    db.add_subnet_range(bd_id, cidr)
                    added += 1
            except ValueError as e:
                errors.append(f"VLAN {vlan_num} row {row_idx}: {e}")
                continue

            # Process IP if present
            if ip_addr:
                try:
                    # Validate IP is within the subnet CIDR range
                    ip_obj = ipaddress.ip_address(ip_addr)
                    cidr_net = ipaddress.ip_network(cidr, strict=False)
                    if ip_obj not in cidr_net:
                        errors.append(f"VLAN {vlan_num} row {row_idx}: IP {ip_addr} is outside subnet range {cidr}")
                        continue

                    ip_id = db.ensure_ip(ip_addr)
                    ip_row = db.get_ip_row(ip_addr)

                    # Link IP to subnet if not already
                    if not ip_row["bd_id"] or ip_row["bd_id"] != bd_id:
                        db.set_ip_links(ip_id, vlan_id, bd_id)
                        added += 1

                    # Process attributes
                    import_keys = set()
                    for col_idx, attr_key in attr_cols:
                        val = sheet.cell(row_idx, col_idx).value
                        if val and not str(val).startswith("[Subnet:"):  # Skip subnet-level markers
                            val_str = str(val).strip()
                            import_keys.add(attr_key)

                            if mode in ("overwrite", "prefer_import"):
                                db.upsert_attr("ip", ip_id, attr_key, val_str, 0)
                                updated += 1
                            elif mode == "prefer_db":
                                existing = db.get_attrs("ip", ip_id)
                                if attr_key not in existing or not existing[attr_key]:
                                    db.upsert_attr("ip", ip_id, attr_key, val_str, 0)
                                    updated += 1
                            elif mode == "interactive":
                                existing = db.get_attrs("ip", ip_id)
                                if attr_key in existing and existing[attr_key] and existing[attr_key] != val_str:
                                    choice = dialog_yes_no(
                                        stdscr,
                                        "Import conflict",
                                        "Attribute conflict",
                                        [
                                            f"IP: {ip_addr}",
                                            f"Key: {attr_key}",
                                            f"DB value: {existing[attr_key]}",
                                            f"Import value: {val_str}",
                                            "",
                                            "Use imported value?"
                                        ],
                                        default_yes=False,
                                        db_name=db_name
                                    )
                                    if choice:
                                        db.upsert_attr("ip", ip_id, attr_key, val_str, 0)
                                        updated += 1
                                else:
                                    db.upsert_attr("ip", ip_id, attr_key, val_str, 0)
                                    updated += 1

                    # Overwrite mode: clear DB attrs not present in the import
                    if mode == "overwrite":
                        existing = db.get_attrs("ip", ip_id)
                        for k, v in existing.items():
                            if k not in import_keys and (v or "").strip():
                                db.delete_attr("ip", ip_id, k)
                                updated += 1

                except Exception as e:
                    errors.append(f"VLAN {vlan_num} row {row_idx} IP {ip_addr}: {e}")

    return added, updated, errors


def workflow_export_all(stdscr, db: DB, db_name: str):
    """Export all VLANs to XLSX file."""
    bc = "Main > Export All"
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        dialog_message(stdscr, bc, "Missing Dependency", ["openpyxl is required for XLSX export.", "", "Install with: pip install openpyxl"], db_name=db_name)
        return

    filename = edit_line_dialog(stdscr, bc, "Export All VLANs", "Filename (blank for auto):", "", db_name=db_name)
    if filename is None:
        return

    filename = filename.strip()
    if not filename:
        filename = ""
    elif not filename.endswith(".xlsx"):
        filename += ".xlsx"

    try:
        result_filename = export_vlans_to_xlsx(db, None, filename)
        dialog_message(stdscr, bc, "Export Complete", [f"Exported to: {result_filename}"], db_name=db_name)
    except Exception as e:
        dialog_message(stdscr, bc, "Export Failed", [f"Error: {e}"], db_name=db_name)


def workflow_export_vlan(stdscr, db: DB, vlan_id: int, db_name: str):
    """Export single VLAN to XLSX file."""
    vlan = db.get_vlan_by_id(vlan_id)
    if not vlan:
        return

    bc = f"VLAN {vlan['vlan_num']} > Export"
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        dialog_message(stdscr, bc, "Missing Dependency", ["openpyxl is required for XLSX export.", "", "Install with: pip install openpyxl"], db_name=db_name)
        return

    filename = edit_line_dialog(stdscr, bc, "Export VLAN", "Filename (blank for auto):", "", db_name=db_name)
    if filename is None:
        return

    filename = filename.strip()
    if not filename:
        filename = ""
    elif not filename.endswith(".xlsx"):
        filename += ".xlsx"

    try:
        result_filename = export_vlans_to_xlsx(db, [vlan_id], filename)
        dialog_message(stdscr, bc, "Export Complete", [f"Exported to: {result_filename}"], db_name=db_name)
    except Exception as e:
        dialog_message(stdscr, bc, "Export Failed", [f"Error: {e}"], db_name=db_name)


def choose_file(stdscr, breadcrumb: str, title: str, extension: str, db_name: str) -> Optional[str]:
    """Simple file chooser - lists files in current directory matching extension."""
    import os

    try:
        files = [f for f in os.listdir('.') if f.endswith(extension)]
        files.sort()
    except Exception:
        files = []

    if not files:
        dialog_message(stdscr, breadcrumb, "No files", [f"No {extension} files found in current directory."], db_name=db_name)
        return None

    rows = [ListRow(f, "", "") for f in files]
    idx, _ = full_screen_list(stdscr, breadcrumb, title, rows, db_name=db_name)

    if idx is None:
        return None

    return files[idx]


def workflow_import(stdscr, db: DB, db_name: str):
    """Import from XLSX file."""
    bc = "Main > Import"
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        dialog_message(stdscr, bc, "Missing Dependency", ["openpyxl is required for XLSX import.", "", "Install with: pip install openpyxl"], db_name=db_name)
        return

    filename = choose_file(stdscr, bc, "Select XLSX file to import", ".xlsx", db_name=db_name)
    if not filename:
        return

    # Select mode
    mode_rows = [
        ListRow("Interactive - prompt for each conflict", "", ""),
        ListRow("Prefer DB - keep existing, add new", "", ""),
        ListRow("Prefer Import - use imported values", "", ""),
        ListRow("Overwrite - replace all attrs from import", "", ""),
    ]

    idx, _ = full_screen_list(stdscr, bc, "Select import mode", mode_rows, db_name=db_name)
    if idx is None:
        return

    modes = ["interactive", "prefer_db", "prefer_import", "overwrite"]
    mode = modes[idx]

    try:
        added, updated, errors = import_from_xlsx(stdscr, db, filename, mode, db_name)

        result_lines = [
            f"Import complete:",
            f"Added: {added}",
            f"Updated: {updated}",
            f"Errors: {len(errors)}",
        ]

        if errors:
            result_lines.append("")
            result_lines.append("Errors:")
            result_lines.extend(errors[:10])  # Show first 10 errors
            if len(errors) > 10:
                result_lines.append(f"... and {len(errors) - 10} more")

        dialog_message(stdscr, bc, "Import Results", result_lines, db_name=db_name)
    except Exception as e:
        dialog_message(stdscr, bc, "Import Failed", [f"Error: {e}"], db_name=db_name)


# =============================================================================
# Audit Log Workflow
# =============================================================================


def workflow_audit_log(stdscr, db: DB, db_name: str):
    """View audit log with filter/sort toggles and rollback."""
    bc = "Main > Audit Log"
    filter_restorable = False  # False = all, True = restorable only
    sort_newest = True  # True = newest first, False = oldest first
    sel = 0
    top = 0

    while True:
        all_entries = db.list_audit_log()  # Always newest-first from DB

        if not all_entries:
            dialog_message(stdscr, bc, "No Audit Log", ["No audit log entries found."], db_name=db_name)
            return

        # Apply filter
        if filter_restorable:
            entries = [e for e in all_entries if e["snapshot_id"]]
        else:
            entries = list(all_entries)

        # Apply sort
        if not sort_newest:
            entries = list(reversed(entries))

        # Build display rows
        rows = []
        for entry in entries:
            timestamp = entry["timestamp"][:19]
            has_snapshot = "✓" if entry["snapshot_id"] else " "
            logged_by = entry["logged_by"] or "-"
            label = f"{timestamp} {has_snapshot} [{logged_by:12}] {entry['action']}: {entry['description']}"
            rows.append(ListRow(label, "", ""))

        filter_txt = "Restorable" if filter_restorable else "All"
        sort_txt = "Newest first" if sort_newest else "Oldest first"
        title = f"Audit Log ({filter_txt}, {sort_txt}) ✓ = has snapshot"
        footer = "Enter: rollback f: filter s: sort q: back Esc: main menu"

        if not rows:
            dialog_message(stdscr, bc, "No Entries", ["No restorable entries found.", "Press 'f' to show all entries."], db_name=db_name)
            filter_restorable = False
            continue

        sel = max(0, min(sel, len(rows) - 1))

        draw_chrome(stdscr, bc, footer, db_name=db_name, db=db)
        bw = body_win(stdscr)
        h, _ = bw.getmaxyx()
        view_h = max(1, h - 5)
        if sel < top:
            top = sel
        if sel >= top + view_h:
            top = sel - view_h + 1

        render_list_rows(bw, title, rows, sel, top)
        bw.refresh()

        c = stdscr.getch()
        if c == 27:
            raise GoHome()
        if c in (ord("q"), ord("Q")):
            return

        if c in (curses.KEY_UP, ord("k")):
            sel = max(0, sel - 1)
        elif c in (curses.KEY_DOWN, ord("j")):
            sel = min(len(rows) - 1, sel + 1)
        elif c == curses.KEY_PPAGE:
            sel = max(0, sel - view_h)
        elif c == curses.KEY_NPAGE:
            sel = min(len(rows) - 1, sel + view_h)
        elif c == curses.KEY_HOME:
            sel = 0
        elif c == curses.KEY_END:
            sel = max(0, len(rows) - 1)

        elif c in (ord("f"), ord("F")):
            filter_restorable = not filter_restorable
            sel = 0
            top = 0

        elif c in (ord("s"), ord("S")):
            sort_newest = not sort_newest
            sel = 0
            top = 0

        elif c in (ord("\n"), ord("\r"), curses.KEY_ENTER):
            entry = entries[sel]

            if not entry["snapshot_id"]:
                dialog_message(
                    stdscr, bc, "No Snapshot",
                    ["This entry has no snapshot.", "Cannot rollback to this point."],
                    db_name=db_name
                )
                continue

            if not require_edit(stdscr, db, bc, "restore snapshots"):
                continue

            confirm = dialog_yes_no(
                stdscr, bc, "Confirm Rollback",
                [
                    f"Restore database to: {entry['timestamp'][:19]}",
                    f"Action: {entry['description']}",
                    "",
                    "WARNING: All changes after this point will be lost!",
                    "A backup will be created before rollback.",
                    "",
                    "Continue with rollback?"
                ],
                default_yes=False, db_name=db_name
            )

            if not confirm:
                continue

            try:
                db_path = sys.argv[1]
                backup_file = db.restore_snapshot(entry["snapshot_id"], db_path)
                dialog_message(
                    stdscr, bc, "Rollback Complete",
                    [
                        f"Database restored to {entry['timestamp'][:19]}",
                        f"Backup saved to: {backup_file}",
                        "",
                        "Press any key to return to main menu."
                    ],
                    db_name=db_name
                )
                return
            except Exception as e:
                dialog_message(stdscr, bc, "Rollback Failed", [f"Error: {e}"], db_name=db_name)


# =============================================================================
# Main menu
# =============================================================================


def mainmenu(stdscr, db: DB, user: User):
    sel = 0
    db_name = db.db_name
    QUIT = "QUIT"

    while True:
        bc = "Main"
        db_name = db.db_name  # Refresh in case configure changed it

        # Build menu as (label, handler) pairs
        if user.is_viewer():
            menu = [
                ("Search", lambda: workflow_search(stdscr, db, db_name=db_name)),
                ("List", lambda: workflow_list(stdscr, db, db_name=db_name)),
                ("Export All VLANs", lambda: workflow_export_all(stdscr, db, db_name=db_name)),
                ("Audit Log", lambda: workflow_audit_log(stdscr, db, db_name=db_name)),
                ("Configure", lambda: workflow_configure(stdscr, db, user, db_name=db_name)),
                ("Quit", QUIT),
            ]
            footer = "Up/Down: move Enter: select q/Esc: jump to Quit [READ-ONLY]"
        else:
            menu = [
                ("Create VLAN", lambda: workflow_create_vlan(stdscr, db, db_name=db_name)),
                ("Create Subnet", lambda: workflow_create_subnet(stdscr, db, db_name=db_name)),
                ("Search", lambda: workflow_search(stdscr, db, db_name=db_name)),
                ("List", lambda: workflow_list(stdscr, db, db_name=db_name)),
                ("Export All VLANs", lambda: workflow_export_all(stdscr, db, db_name=db_name)),
                ("Import from XLSX", lambda: workflow_import(stdscr, db, db_name=db_name)),
                ("Audit Log", lambda: workflow_audit_log(stdscr, db, db_name=db_name)),
                ("Configure", lambda: workflow_configure(stdscr, db, user, db_name=db_name)),
                ("Quit", QUIT),
            ]
            footer = "Up/Down: move Enter: select q/Esc: jump to Quit"

        rows = [ListRow(label, "", "") for label, _ in menu]
        quit_idx = len(menu) - 1

        try:
            idx, sel = full_screen_list(
                stdscr, bc, "Main menu", rows,
                db_name=db_name, footer=footer, start_sel=sel, db=db,
            )
        except GoHome:
            sel = quit_idx
            continue

        if idx is None:
            sel = quit_idx
            continue

        _, action = menu[idx]
        if action is QUIT:
            return
        action()


# =============================================================================
# BIOS Boot Sequence (Easter Egg)
# =============================================================================

# Default BIOS configuration
BIOS_DEFAULTS = {
    'cpu_type': 'Intel 80486DX2-66',
    'base_memory': 640,
    'extended_memory': 3072,
    'floppy_drives': 1,
    'hard_drives': 1,
    'boot_speed': 1.0,
    'post_messages': [
        'Initializing System Hardware...',
        'Testing Base Memory...',
        'Testing Extended Memory...',
        'Detecting Floppy Drives...',
        'Detecting Hard Drives...',
        'Initializing Keyboard Controller...',
        'Setting up System Timer...',
        'Configuring PCI Bus...',
        'Enabling A20 Gate...',
        'Shadow RAM Configuration...',
    ]
}


def bios_get_config(db: DB) -> dict:
    """Load BIOS configuration from database, with defaults."""
    config = BIOS_DEFAULTS.copy()
    config['cpu_type'] = db.get_config('bios_cpu', config['cpu_type'])
    config['base_memory'] = int(db.get_config('bios_base_memory', str(config['base_memory'])))
    config['extended_memory'] = int(db.get_config('bios_extended_memory', str(config['extended_memory'])))
    config['floppy_drives'] = int(db.get_config('bios_floppy_drives', str(config['floppy_drives'])))
    config['hard_drives'] = int(db.get_config('bios_hard_drives', str(config['hard_drives'])))
    config['boot_speed'] = float(db.get_config('bios_boot_speed', str(config['boot_speed'])))
    return config


def bios_delay(stdscr, duration: float, speed_multiplier: float = 1.0) -> bool:
    """Sleep for duration adjusted by speed multiplier, allow ESC to skip."""
    actual_duration = duration / max(0.1, speed_multiplier)
    start_time = time.time()

    stdscr.nodelay(True)
    try:
        while time.time() - start_time < actual_duration:
            key = stdscr.getch()
            if key == 27:  # ESC key
                return True  # Signal to skip
            time.sleep(0.01)
    finally:
        stdscr.nodelay(False)

    return False


def bios_print_line(stdscr, row: int, text: str, col: int = 0, attr: int = 0):
    """Print a line at specified row and column."""
    try:
        h, w = stdscr.getmaxyx()
        if row < h and col < w:
            stdscr.addnstr(row, col, text, w - col - 1, attr)
            stdscr.refresh()
    except curses.error:
        pass


def run_bios_sequence(stdscr, db: DB):
    """Run the BIOS boot sequence animation."""
    config = bios_get_config(db)

    # Initialize colors - green on black for retro feel
    curses.start_color()
    curses.init_pair(10, curses.COLOR_GREEN, curses.COLOR_BLACK)
    bios_color = curses.color_pair(10)

    # Clear screen and set colors
    stdscr.clear()
    stdscr.bkgd(' ', bios_color)
    curses.curs_set(0)

    speed = config['boot_speed']
    skip = False
    h, w = stdscr.getmaxyx()

    # BIOS Header
    row = 0
    header_line = "=" * min(78, w - 2)
    bios_print_line(stdscr, row, header_line, 0, bios_color | curses.A_BOLD)
    row += 1
    bios_print_line(stdscr, row, "  AmiBIOS (C) 1994 American Megatrends Inc.", 0, bios_color | curses.A_BOLD)
    row += 1
    bios_print_line(stdscr, row, "  BIOS Version 1.00.12.AX1T", 0, bios_color)
    row += 1
    bios_print_line(stdscr, row, header_line, 0, bios_color | curses.A_BOLD)
    row += 2

    skip = bios_delay(stdscr, 0.5, speed)

    # CPU Detection
    if not skip:
        bios_print_line(stdscr, row, f"  Processor: {config['cpu_type']}", 0, bios_color)
        row += 1
        skip = bios_delay(stdscr, 0.3, speed)

    # Memory Test - Base Memory
    if not skip:
        row += 1
        memory_row = row

        # Animate memory counting
        steps = 20
        for i in range(steps + 1):
            if skip:
                break
            current = int((config['base_memory'] * i) / steps)
            bios_print_line(stdscr, memory_row, f"  Base Memory: {current:6d} KB       ", 0, bios_color)
            skip = bios_delay(stdscr, 0.02, speed)

        bios_print_line(stdscr, memory_row, f"  Base Memory: {config['base_memory']:6d} KB  OK", 0, bios_color)
        row += 1

    # Extended Memory
    if not skip:
        ext_row = row

        # Animate extended memory counting
        steps = 30
        for i in range(steps + 1):
            if skip:
                break
            current = int((config['extended_memory'] * i) / steps)
            bios_print_line(stdscr, ext_row, f"  Extended Memory: {current:6d} KB       ", 0, bios_color)
            skip = bios_delay(stdscr, 0.015, speed)

        bios_print_line(stdscr, ext_row, f"  Extended Memory: {config['extended_memory']:6d} KB  OK", 0, bios_color)
        row += 2

    # Floppy Drives
    if not skip and config['floppy_drives'] > 0:
        skip = bios_delay(stdscr, 0.2, speed)
        bios_print_line(stdscr, row, "  Floppy Drives:", 0, bios_color)
        row += 1
        for i in range(config['floppy_drives']):
            if skip:
                break
            drive_letter = chr(65 + i)  # A, B, etc.
            bios_print_line(stdscr, row, f"    Drive {drive_letter}: 1.44 MB 3.5\"", 0, bios_color)
            row += 1
            skip = bios_delay(stdscr, 0.15, speed)
        row += 1

    # Hard Drives
    if not skip and config['hard_drives'] > 0:
        skip = bios_delay(stdscr, 0.2, speed)
        bios_print_line(stdscr, row, "  Hard Drives:", 0, bios_color)
        row += 1

        hd_sizes = [540, 1024, 2048, 4096]
        hd_positions = ["Primary Master", "Primary Slave", "Secondary Master", "Secondary Slave"]
        for i in range(min(config['hard_drives'], 4)):
            if skip:
                break
            size = hd_sizes[i % len(hd_sizes)]
            pos = hd_positions[i]
            bios_print_line(stdscr, row, f"    {pos}: {size} MB", 0, bios_color)
            row += 1
            skip = bios_delay(stdscr, 0.2, speed)
        row += 1

    # POST Messages
    if not skip:
        row += 1
        bios_print_line(stdscr, row, "  POST Sequence:", 0, bios_color | curses.A_BOLD)
        row += 1

        for msg in config['post_messages']:
            if skip or row > h - 5:
                break
            bios_print_line(stdscr, row, f"    {msg}", 0, bios_color)
            skip = bios_delay(stdscr, 0.15, speed)

            # Print OK status
            padded_msg = f"    {msg:<50}"
            bios_print_line(stdscr, row, f"{padded_msg} [  OK  ]", 0, bios_color)
            row += 1

    # Boot completion
    if not skip:
        row += 2
        skip = bios_delay(stdscr, 0.5, speed)

    if not skip:
        bios_print_line(stdscr, row, "  System Configuration Complete", 0, bios_color | curses.A_BOLD)
        row += 1
        skip = bios_delay(stdscr, 0.3, speed)

    if not skip:
        bios_print_line(stdscr, row, "  Press any key to continue...", 0, bios_color | curses.A_BLINK)
        stdscr.nodelay(False)
        stdscr.getch()

    # Final screen clear with "booting" message
    stdscr.clear()
    bios_print_line(stdscr, h // 2, "  Loading IPAM / VLAN Manager...", 0, bios_color | curses.A_BOLD)
    stdscr.refresh()
    time.sleep(0.5)


def workflow_bios_settings(stdscr, db: DB, db_name: str):
    """Configure BIOS easter egg settings."""
    bc = "Main > Configure > BIOS Settings"

    while True:
        config = bios_get_config(db)

        options = [
            ListRow(f"CPU Type: {config['cpu_type']}", "", ""),
            ListRow(f"Base Memory: {config['base_memory']} KB", "", ""),
            ListRow(f"Extended Memory: {config['extended_memory']} KB", "", ""),
            ListRow(f"Floppy Drives: {config['floppy_drives']}", "", ""),
            ListRow(f"Hard Drives: {config['hard_drives']}", "", ""),
            ListRow(f"Boot Speed: {config['boot_speed']}x", "", ""),
            ListRow("Test BIOS Sequence", "", ""),
            ListRow("Reset to Defaults", "", ""),
            ListRow("Hide BIOS Settings", "", ""),
            ListRow("Back", "", ""),
        ]

        idx, _ = full_screen_list(stdscr, bc, "BIOS Boot Settings", options, db_name=db_name)
        if idx is None or idx == 9:
            return

        if idx == 0:
            # CPU Type
            cpu_presets = [
                ListRow("Intel 8086", "", ""),
                ListRow("Intel 80286", "", ""),
                ListRow("Intel 80386DX-33", "", ""),
                ListRow("Intel 80486DX2-66", "", ""),
                ListRow("Intel Pentium 75", "", ""),
                ListRow("Intel Pentium 100", "", ""),
                ListRow("Intel Pentium 133", "", ""),
                ListRow("AMD Am486DX4-120", "", ""),
                ListRow("Cyrix 6x86 PR166+", "", ""),
                ListRow("Custom...", "", ""),
            ]
            cpu_idx, _ = full_screen_list(stdscr, bc, "Select CPU Type", cpu_presets, db_name=db_name)
            if cpu_idx is not None:
                if cpu_idx == 9:
                    # Custom
                    custom_cpu = edit_line_dialog(stdscr, bc, "CPU Type", "Enter CPU name:", config['cpu_type'], db_name=db_name)
                    if custom_cpu:
                        db.set_config('bios_cpu', custom_cpu.strip())
                else:
                    db.set_config('bios_cpu', cpu_presets[cpu_idx].label)

        elif idx == 1:
            # Base Memory
            mem = edit_line_dialog(stdscr, bc, "Base Memory", "Base memory in KB (typical: 640):", str(config['base_memory']), db_name=db_name)
            if mem and mem.strip().isdigit():
                db.set_config('bios_base_memory', mem.strip())

        elif idx == 2:
            # Extended Memory
            mem = edit_line_dialog(stdscr, bc, "Extended Memory", "Extended memory in KB:", str(config['extended_memory']), db_name=db_name)
            if mem and mem.strip().isdigit():
                db.set_config('bios_extended_memory', mem.strip())

        elif idx == 3:
            # Floppy Drives
            floppy_options = [ListRow(f"{i} drive(s)", "", "") for i in range(5)]
            fl_idx, _ = full_screen_list(stdscr, bc, "Number of Floppy Drives", floppy_options, db_name=db_name)
            if fl_idx is not None:
                db.set_config('bios_floppy_drives', str(fl_idx))

        elif idx == 4:
            # Hard Drives
            hd_options = [ListRow(f"{i} drive(s)", "", "") for i in range(5)]
            hd_idx, _ = full_screen_list(stdscr, bc, "Number of Hard Drives", hd_options, db_name=db_name)
            if hd_idx is not None:
                db.set_config('bios_hard_drives', str(hd_idx))

        elif idx == 5:
            # Boot Speed
            speed_options = [
                ListRow("0.5x (Slow)", "", ""),
                ListRow("1.0x (Normal)", "", ""),
                ListRow("1.5x (Fast)", "", ""),
                ListRow("2.0x (Very Fast)", "", ""),
                ListRow("5.0x (Ludicrous)", "", ""),
            ]
            sp_idx, _ = full_screen_list(stdscr, bc, "Boot Speed", speed_options, db_name=db_name)
            if sp_idx is not None:
                speeds = ["0.5", "1.0", "1.5", "2.0", "5.0"]
                db.set_config('bios_boot_speed', speeds[sp_idx])

        elif idx == 6:
            # Test BIOS Sequence
            run_bios_sequence(stdscr, db)
            init_theme(stdscr, db)  # Restore normal theme

        elif idx == 7:
            # Reset to Defaults
            confirm = dialog_yes_no(stdscr, bc, "Reset BIOS Settings", ["Reset all BIOS settings to defaults?"], default_yes=False, db_name=db_name)
            if confirm:
                for key in ['bios_cpu', 'bios_base_memory', 'bios_extended_memory',
                           'bios_floppy_drives', 'bios_hard_drives', 'bios_boot_speed']:
                    db.x("DELETE FROM config WHERE key=?", (key,))
                dialog_message(stdscr, bc, "Reset Complete", ["BIOS settings reset to defaults."], db_name=db_name)

        elif idx == 8:
            # Hide BIOS Settings
            confirm = dialog_yes_no(
                stdscr, bc, "Hide BIOS Settings",
                ["Hide BIOS settings from Configure menu?", "", "Run with --bios flag to reveal again."],
                default_yes=False, db_name=db_name
            )
            if confirm:
                db.set_config('bios_unlocked', '0')
                dialog_message(stdscr, bc, "Hidden", ["BIOS settings hidden.", "Use --bios flag to reveal."], db_name=db_name)
                return


def main():
    if "--version" in sys.argv:
        print(f"IPAM/VLAN Manager TUI v{VERSION}")
        sys.exit(0)

    if len(sys.argv) < 2:
        print(f"Usage: {sys.argv[0]} path/to/ipam.db [--bios] [--reset-colors] [--reset-admin] [--version]")
        sys.exit(2)

    # Parse flags
    show_bios = "--bios" in sys.argv
    reset_colors = "--reset-colors" in sys.argv
    reset_admin = "--reset-admin" in sys.argv

    # Get database path (first non-flag argument)
    db_path = None
    for arg in sys.argv[1:]:
        if not arg.startswith("--"):
            db_path = arg
            break

    if not db_path:
        print(f"Usage: {sys.argv[0]} path/to/ipam.db [--bios] [--reset-colors] [--reset-admin] [--version]")
        sys.exit(2)

    db = DB(db_path)
    db.init()

    # Check for --reset-colors flag
    if reset_colors:
        db.set_config("fg_color", "green")
        print("Foreground color reset to green (default)")
        db.close()
        return

    # Check for --reset-admin flag (resets admin password to 'admin')
    if reset_admin:
        admin_user = db.get_user_by_username("admin")
        if admin_user:
            db.update_user_password(admin_user.id, "admin")
            print("Admin password reset to 'admin'")
        else:
            # Create admin user if it doesn't exist
            db.create_user("admin", "admin", ROLE_ADMIN)
            print("Admin user created with password 'admin'")
        db.close()
        return

    # Handle --bios flag: unlock settings and show BIOS sequence
    if show_bios:
        db.set_config('bios_unlocked', '1')

    db_name = db.get_config("db_name", "").strip()
    db.db_name = db_name
    current_user: Optional[User] = None

    def run(stdscr):
        nonlocal current_user

        # Login screen
        current_user = login_screen(stdscr, db)
        if current_user is None:
            return  # User cancelled login

        # Show BIOS boot sequence if flag was passed
        if show_bios:
            run_bios_sequence(stdscr, db)

        init_theme(stdscr, db, current_user)
        while True:
            try:
                mainmenu(stdscr, db, current_user)
                return
            except GoHome:
                continue

    try:
        curses.wrapper(run)
    finally:
        if current_user:
            db.log_action("logout", f"User '{current_user.username}' logged out")
        db.close()


if __name__ == "__main__":
    main()
